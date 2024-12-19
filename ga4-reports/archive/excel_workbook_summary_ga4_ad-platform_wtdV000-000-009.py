# file name: ga4-reports/excel_workbook_summary_ga4_ad-platform_wtd.py
# version: V000-000-009
# output:
# Notes: improves the formatting of each worksheet

import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill

# Load the CSV into a DataFrame
df = pd.read_csv('ga4-reports/output/ga4_summary_metrics_multiple_periods.csv')

# Rename columns with '_Actual' suffix
columns_to_add_actual_suffix = [
    'FF_Lead_Event_Count', 'FF_Purchase_Event_Count', 'FF_BRSubmit_Event_Count',
    'FF_DMSubmit_Event_Count', 'FF_PhoneGet_Event_Count', 'Housing Requests',
    'Total Traveler Actions', 'Traveler Value', 'Landlord Value', 'Sessions',
    'Impressions', 'Clicks', 'Cost', 'Lead_Conversion', 'CPL', 'CAC',
    'Traveler_Conversion', 'CPTA', 'ROAS'
]
df.rename(columns={col: f'{col}_Actual' for col in columns_to_add_actual_suffix}, inplace=True)

# List of periods
periods = ['WTD', 'EOW', 'MTD', 'QTD']

# Create a new Excel workbook
wb = openpyxl.Workbook()

# Iterate over each period
for period in periods:
    # Filter the DataFrame for the current period
    period_df = df[df['Period'] == period]

    # Determine the start and end dates
    start_date = period_df['Start_Date_Period'].min()
    end_date = period_df['End_Date_Period'].max()

    # Drop the unnecessary columns
    period_df = period_df.drop(columns=['Start_Date_Period', 'End_Date_Period', 'Period'])

    # Add a new worksheet for the period
    ws = wb.create_sheet(title=period)

    # Set the title and sub-title
    ws['A1'] = f"GROWTH - PERFORMANCE REPORT - {period}"
    ws['A1'].font = Font(name="Aptos Display", size=18, bold=True, color="FFFFFF")
    ws['A1'].alignment = Alignment(horizontal="left")

    ws['A2'] = f"{start_date} to {end_date}"
    ws['A2'].font = Font(name="Aptos Display", size=16, bold=True, color="FFFFFF")
    ws['A2'].alignment = Alignment(horizontal="left")

    # Insert the DataFrame into the worksheet
    for r in dataframe_to_rows(period_df, index=False, header=True):
        ws.append(r)

    # Split metric names into two rows
    metric_names = [col.split("_")[0] for col in period_df.columns if col not in ['Channel_Group', 'Campaign_Group']]
    unique_metric_names = sorted(set(metric_names), key=metric_names.index)

    # Insert a new row above the current column headers for the metric names
    ws.insert_rows(5)

    # Column offset for merging metric names across four columns
    col_offset = 2

    for idx, metric_name in enumerate(unique_metric_names):
        start_col = idx * 4 + col_offset + 1
        end_col = start_col + 3

        # Set the metric name in row 5 and merge cells
        ws.cell(row=5, column=start_col, value=metric_name)
        ws.merge_cells(start_row=5, start_column=start_col, end_row=5, end_column=end_col)
        ws.cell(row=5, column=start_col).alignment = Alignment(horizontal="center")

        # Set the period labels in row 6
        period_labels = ["Actual", "LW", "WoW", "YoY"]
        for i, period_label in enumerate(period_labels):
            ws.cell(row=6, column=start_col + i, value=period_label)
            ws.cell(row=6, column=start_col + i).alignment = Alignment(horizontal="center")

    # Apply formatting logic
    base_font = Font(name="Aptos Display", size=12)
    header_font = Font(name="Aptos Display", size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="024991", end_color="024991", fill_type="solid")

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = base_font

    for col in ws.iter_cols(min_col=3, max_col=ws.max_column, min_row=5, max_row=6):
        for cell in col:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    # Apply alternating row colors
    alt_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    for row in range(7, ws.max_row + 1):
        if (row - 7) % 2 == 0:
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    cell.fill = alt_fill

    # Adjust column width for A & B
    for col in ['A', 'B']:
        max_length = max([len(str(cell.value)) for cell in ws[col] if cell.value])
        ws.column_dimensions[col].width = max_length * 1.2

    # Apply freeze panes
    ws.freeze_panes = 'C7'

# Remove the default "Sheet" created by openpyxl
del wb['Sheet']

# Save the workbook
wb.save('ga4-reports/output/ga4_summary_metrics_formatted.xlsx')

print("Formatted Excel file with multiple sheets has been created and saved.")
