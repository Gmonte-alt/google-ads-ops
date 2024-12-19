# file name:
# version: V000-000-002
# output:
# notes: Builds on V000-000-001 but outputs and places columns in metric : mtd calculation column-hierarchy format. 

import pandas as pd
from datetime import datetime, timedelta

# Load CSV data
data = pd.read_csv('ga4-reports/output/ga4_combined_data.csv') # 'mtd_performance.csv')

# Automate the calculation of the reference_date as the day before yesterday
reference_date = datetime.now()
reference_date = reference_date.replace(hour=0, minute=0, second=0, microsecond=0)
end_date = reference_date - timedelta(days=2)
print(reference_date)
print(end_date)

# Convert Date column from integer format YYYYMMDD to datetime
data['Date'] = pd.to_datetime(data['Date'].astype(str), format='%Y%m%d')

# Filter data up to the reference date (two days ago) and within the current month
data['Date'] = pd.to_datetime(data['Date'])

start_of_month = end_date.replace(day=1) # changing to end date: reference_date.replace(day=1)
data = data[(data['Date'] <= end_date) & (data['Date'] >= start_of_month)]
print(start_of_month)

# Calculate the number of days passed in the current month
days_passed = (end_date - start_of_month).days + 1  # Including the current day
print(days_passed)

# Calculate days left in the month
days_in_month = (reference_date.replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)
days_left = (days_in_month - reference_date).days + 1

# Filter data to only include rows where Channel_Group contains "SEM"
data = data[data['Channel_Group'].str.contains('SEM')]

# Rename columns in the DataFrame
data.rename(columns={
    'Cost': 'Spend',
    'FF_Purchase_Event_Count': 'New Properties'
}, inplace=True)

# Calculate 'Tenant Actions' by summing the three relevant metrics
data['Tenant Actions'] = (
    data['FF_BRSubmit_Event_Count'] + 
    data['FF_DMSubmit_Event_Count'] + 
    data['FF_PhoneGet_Event_Count']
)

# Define the budget for each combination of Channel_Group and Campaign_Group
budget_dict = {
    ('SEM Brand', 'Google Search - Brand'): {'Spend': 74913, 'New Properties': 2618, 'Tenant Actions': 65202},
    ('SEM Brand', 'Google Search - Brand Broad'): {'Spend': 68946, 'New Properties': 255, 'Tenant Actions': 8761},
    ('SEM Non-Brand', 'Google LL Non-Brand'): {'Spend': 78224, 'New Properties': 446, 'Tenant Actions': 1743},
    ('SEM Non-Brand - Tenant', 'Google Tenant'): {'Spend': 133997, 'New Properties': 88, 'Tenant Actions': 18882},
    ('SEM Brand', 'Bing Brand'): {'Spend': 30000, 'New Properties': 682, 'Tenant Actions': 19341},
    ('SEM Non-Brand', 'Bing LL Non-Brand'): {'Spend': 16682, 'New Properties': 70, 'Tenant Actions': 981},
    ('SEM Brand', 'Google TNH Brand'): {'Spend': 64000, 'New Properties': 303, 'Tenant Actions': 1339},
    ('SEM Non-Brand', 'Google TNH Non-Brand'): {'Spend': 128000, 'New Properties': 607, 'Tenant Actions': 2261}
}

# Initialize dictionaries for MTD metrics and budgets
mtd_metrics = {}
mtd_budget = {}
ratios = {}
mtd_ratios = {}

# Calculate the number of days passed in the current month
days_passed = (end_date - start_of_month).days + 1  # Including the current day

# Filter and aggregate data based on Channel_Group and Campaign_Group
for (channel_group, campaign_group_name), budgets in budget_dict.items():
    if "Google LL Non-Brand" in campaign_group_name:
        filtered_data = data[(data['Channel_Group'] == channel_group) & 
                             (data['Campaign_Group'].str.contains('Google') & ~data['Campaign_Group'].str.contains('TNH'))]
    elif "Google Tenant" in campaign_group_name:
        filtered_data = data[(data['Channel_Group'] == channel_group) & 
                             (data['Campaign_Group'].str.contains('Google') & ~data['Campaign_Group'].str.contains('TNH'))]
    elif "Bing Brand" in campaign_group_name:
        filtered_data = data[(data['Channel_Group'] == channel_group) & 
                             (data['Campaign_Group'].str.contains('Bing'))]
    elif "Bing LL Non-Brand" in campaign_group_name:
        filtered_data = data[(data['Channel_Group'] == channel_group) & 
                             (data['Campaign_Group'].str.contains('Bing'))]
    elif "Google TNH Brand" in campaign_group_name:
        filtered_data = data[(data['Channel_Group'] == channel_group) & 
                             (data['Campaign_Group'].str.contains('TNH'))]
    elif "Google TNH Non-Brand" in campaign_group_name:
        filtered_data = data[(data['Channel_Group'] == channel_group) & 
                             (data['Campaign_Group'].str.contains('TNH'))]
    else:
        filtered_data = data[(data['Channel_Group'] == channel_group) & 
                             (data['Campaign_Group'] == campaign_group_name)]

    # Aggregate MTD performance for the filtered data
    mtd_performance = filtered_data.agg({
        'Spend': 'sum',
        'New Properties': 'sum',
        'Tenant Actions': 'sum'
    })
    
    # Calculate MTD metrics
    mtd_metrics[(channel_group, campaign_group_name)] = mtd_performance.to_dict()
    
    # Calculate the MTD budget by prorating the monthly budget
    mtd_budget[(channel_group, campaign_group_name)] = {
        metric: (budget / days_in_month.day) * days_passed
        for metric, budget in budgets.items()
    }
    
    # Calculate the ratio of MTD to total budget
    ratios[(channel_group, campaign_group_name)] = {
        metric: mtd_metrics[(channel_group, campaign_group_name)][metric] / budgets[metric]
        for metric in mtd_metrics[(channel_group, campaign_group_name)]
    }
    
    # Calculate the ratio of MTD to MTD budget
    mtd_ratios[(channel_group, campaign_group_name)] = {
        metric: mtd_metrics[(channel_group, campaign_group_name)][metric] / mtd_budget[(channel_group, campaign_group_name)][metric]
        for metric in mtd_metrics[(channel_group, campaign_group_name)]
    }

# Compile the final report
report = {
    'MTD Metrics': mtd_metrics,
    'MTD Budget': mtd_budget,
    'Ratio of MTD to MTD Budget': mtd_ratios,
    'Ratio of MTD to Total Budget': ratios,
}

# ------------------------------------------------------------------------------------------- #
# ----------------------------- EXCEL WORKBOOK BUILD ---------------------------------------- #
# ------------------------------------------------------------------------------------------- #

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# Load the Excel template
template_path = 'budgetting/input/MTD_Budget_Output_Excel_Template.xlsx'
workbook = load_workbook(template_path)

# Create a new worksheet for the report
sheet_name = reference_date.strftime('%b') + "_MTD_Budget"
worksheet = workbook.create_sheet(title=sheet_name)

# Write the headers to the worksheet
metrics = ['Spend', 'New Properties', 'Tenant Actions']
column_start = 3  # Start writing metrics in the third column

for i, metric in enumerate(metrics):
    worksheet.merge_cells(start_row=1, start_column=column_start + 3*i, end_row=1, end_column=column_start + 3*i + 2)
    worksheet.cell(row=1, column=column_start + 3*i, value=metric)
    worksheet.cell(row=2, column=column_start + 3*i, value='MTD')
    worksheet.cell(row=2, column=column_start + 3*i + 1, value='MTD Budget')
    worksheet.cell(row=2, column=column_start + 3*i + 2, value='vs Budget')

# Initialize row_num after headers and Total row
row_num = 4

# Write the report data to the worksheet
for (channel_group, campaign_group_name), metric_values in mtd_metrics.items():
    worksheet.cell(row=row_num, column=1, value=channel_group)
    worksheet.cell(row=row_num, column=2, value=campaign_group_name)
    for i, metric in enumerate(metrics):
        worksheet.cell(row=row_num, column=column_start + 3*i, value=metric_values[metric])
        worksheet.cell(row=row_num, column=column_start + 3*i + 1, value=mtd_budget[(channel_group, campaign_group_name)][metric])
        worksheet.cell(row=row_num, column=column_start + 3*i + 2, value=mtd_ratios[(channel_group, campaign_group_name)][metric])
    row_num += 1

# Insert Total row and set up summation formulas
worksheet.cell(row=3, column=1, value="Total")
last_data_row = row_num - 1  # Calculate the last data row number
for i in range(len(metrics)):
    col_base = column_start + 3*i
    worksheet.cell(row=3, column=col_base, value=f"=SUM({worksheet.cell(row=4, column=col_base).coordinate}:{worksheet.cell(row=last_data_row, column=col_base).coordinate})")
    worksheet.cell(row=3, column=col_base + 1, value=f"=SUM({worksheet.cell(row=4, column=col_base + 1).coordinate}:{worksheet.cell(row=last_data_row, column=col_base + 1).coordinate})")
    worksheet.cell(row=3, column=col_base + 2, value=f"={worksheet.cell(row=3, column=col_base).coordinate}/{worksheet.cell(row=3, column=col_base + 1).coordinate}")

# Apply formatting to MTD, MTD Budget, and vs Budget columns
for row in range(3, row_num):  # Start from the first data row
    for i, metric in enumerate(metrics):
        col_base = column_start + 3*i
        # Currency format for Spend and numeric format for New Properties and Tenant Actions
        if metric == 'Spend':
            worksheet.cell(row=row, column=col_base).number_format = '$#,##0'
            worksheet.cell(row=row, column=col_base + 1).number_format = '$#,##0'
        else:
            worksheet.cell(row=row, column=col_base).number_format = '#,##0'
            worksheet.cell(row=row, column=col_base + 1).number_format = '#,##0'
        
        # Percentage format for vs Budget
        worksheet.cell(row=row, column=col_base + 2).number_format = '0%'

# Center align headers and adjust column width
for col in range(1, column_start + 3*len(metrics) + 1):  # Include all columns
    worksheet.cell(row=1, column=col).alignment = worksheet.cell(row=2, column=col).alignment = Alignment(horizontal='center')
    worksheet.column_dimensions[get_column_letter(col)].width = 15  # Adjust width as necessary

# Save the workbook
output_path = 'budgetting/output/MTD_Budget_Output.xlsx'
workbook.save(output_path)

# Display the report
for key, values in report.items():
    print(f"\n{key}:")
    for (channel_group, campaign_group_name), metrics in values.items():
        print(f"  Channel Group: {channel_group}, Campaign Group: {campaign_group_name}")
        for metric, value in metrics.items():
            print(f"    {metric}: {value:.2f}")

print(f"\nThe report has been saved to {output_path}.")