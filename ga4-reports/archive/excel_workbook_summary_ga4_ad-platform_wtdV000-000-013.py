# file name: ga4-reports/excel_workbook_summary_ga4_ad-platform_wtd.py
# version: V000-000-013
# output:
# Notes: Builds on V000-000-012, formats the final output of the EOW Weekly Commentary

import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, NamedStyle, PatternFill, Border, Side
from openpyxl import load_workbook
import gspread
from google.oauth2.service_account import Credentials
from datetime import datetime
import isoweek

# -------- Google Sheets Integration -------- #

# Path to your service account JSON key file
SERVICE_ACCOUNT_FILE = 'insert secret'

# Define the scopes required for the Google Sheets API
SCOPES = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive']

# Create a credentials object using the service account file
creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)

# Authenticate and initialize the gspread client
client = gspread.authorize(creds)

# Open the Google Sheet by its name or URL
sheet = client.open('FF - Growth - Weekly Commentary').sheet1

# Get all rows from the Google Sheet
all_rows = sheet.get_all_values()
headers = all_rows[0]  # First row as headers
data = all_rows[1:]  # Data rows

# -------- Excel Workbook Creation and Formatting -------- #

excel_file_path = 'ga4-reports/output/ga4_summary_metrics_formatted.xlsx'

# Load the CSV into a DataFrame
df = pd.read_csv('ga4-reports/output/ga4_summary_metrics_multiple_periods.csv')

# List of periods
periods = ['WTD', 'EOW', 'MTD', 'QTD']

# Create a new Excel workbook
wb = openpyxl.Workbook()

# Set the base font style and size across the entire worksheet
base_font = Font(name="Aptos Display", size=12)

# Define or get existing formatting styles
dollar_no_decimal = NamedStyle(name="dollar_no_decimal", font=base_font)
dollar_no_decimal.number_format = '"$"#,##0'
wb.add_named_style(dollar_no_decimal)

dollar_two_decimal = NamedStyle(name="dollar_two_decimal", font=base_font)
dollar_two_decimal.number_format = '"$"#,##0.00'
wb.add_named_style(dollar_two_decimal)

percent_no_decimal = NamedStyle(name="percent_no_decimal", font=base_font)
percent_no_decimal.number_format = '0%'
wb.add_named_style(percent_no_decimal)

percent_two_decimal = NamedStyle(name="percent_two_decimal", font=base_font)
percent_two_decimal.number_format = '0.00%'
wb.add_named_style(percent_two_decimal)

comma_no_decimal = NamedStyle(name="comma_no_decimal", font=base_font)
comma_no_decimal.number_format = '#,##0'
wb.add_named_style(comma_no_decimal)


# Iterate over each period
for period in periods:
    # Filter the DataFrame for the current period
    period_df = df[df['Period'] == period]

    # Determine the start and end dates
    start_date = period_df['Start_Date_Period'].min()
    end_date = period_df['End_Date_Period'].max()

    # Drop the unnecessary columns
    period_df = period_df.drop(columns=['Start_Date_Period', 'End_Date_Period', 'Period'])

    # Rename columns according to the mapping
    rename_map = {
        'FF_Lead_Event_Count': 'Leads_Actual',
        'FF_Purchase_Event_Count': 'New Properties_Actual',
        'FF_BRSubmit_Event_Count': 'Booking Requests_Actual',
        'FF_DMSubmit_Event_Count': 'Direct Messages_Actual',
        'FF_PhoneGet_Event_Count': 'Phone Reveals_Actual',
        'Housing Requests': 'Housing Requests_Actual',
        'Total Traveler Actions': 'Total Traveler Actons_Actual',
        'Traveler Value': 'Traveler Value_Actual',
        'Landlord Value': 'Landlord Value_Actual',
        'Sessions': 'Sessions_Actual',
        'Impressions': 'Impressions_Actual',
        'Clicks': 'Clicks_Actual',
        'Cost': 'Spend_Actual',
        'Lead_Conversion': 'Lead Conversion_Actual',
        'CPL': 'CPL_Actual',
        'CAC': 'CAC_Actual',
        'Traveler_Conversion': 'Traveler Conversion_Actual',
        'CPTA': 'CPTA_Actual',
        'ROAS': 'ROAS_Actual',
        'FF_Lead_Event_Count_previous': 'Leads_LW',
        'FF_Purchase_Event_Count_previous': 'New Properties_LW',
        'FF_BRSubmit_Event_Count_previous': 'Booking Requests_LW',
        'FF_DMSubmit_Event_Count_previous': 'Direct Messages_LW',
        'FF_PhoneGet_Event_Count_previous': 'Phone Reveals_LW',
        'Housing Requests_previous': 'Housing Requests_LW',
        'Total Traveler Actions_previous': 'Total Traveler Actons_LW',
        'Traveler Value_previous': 'Traveler Value_LW',
        'Landlord Value_previous': 'Landlord Value_LW',
        'Sessions_previous': 'Sessions_LW',
        'Impressions_previous': 'Impressions_LW',
        'Clicks_previous': 'Clicks_LW',
        'Cost_previous': 'Spend_LW',
        'Lead_Conversion_previous': 'Lead Conversion_LW',
        'CPL_previous': 'CPL_LW',
        'CAC_previous': 'CAC_LW',
        'Traveler_Conversion_previous': 'Traveler Conversion_LW',
        'CPTA_previous': 'CPTA_LW',
        'ROAS_previous': 'ROAS_LW',
        'FF_Lead_Event_Count_WoW': 'Leads_WoW',
        'FF_Purchase_Event_Count_WoW': 'New Properties_WoW',
        'FF_BRSubmit_Event_Count_WoW': 'Booking Requests_WoW',
        'FF_DMSubmit_Event_Count_WoW': 'Direct Messages_WoW',
        'FF_PhoneGet_Event_Count_WoW': 'Phone Reveals_WoW',
        'Housing Requests_WoW': 'Housing Requests_WoW',
        'Total Traveler Actions_WoW': 'Total Traveler Actons_WoW',
        'Traveler Value_WoW': 'Traveler Value_WoW',
        'Landlord Value_WoW': 'Landlord Value_WoW',
        'Sessions_WoW': 'Sessions_WoW',
        'Impressions_WoW': 'Impressions_WoW',
        'Clicks_WoW': 'Clicks_WoW',
        'Cost_WoW': 'Spend_WoW',
        'Lead_Conversion_WoW': 'Lead Conversion_WoW',
        'CPL_WoW': 'CPL_WoW',
        'CAC_WoW': 'CAC_WoW',
        'Traveler_Conversion_WoW': 'Traveler Conversion_WoW',
        'CPTA_WoW': 'CPTA_WoW',
        'ROAS_WoW': 'ROAS_WoW',
        'FF_Lead_Event_Count_YoY': 'Leads_YoY',
        'FF_Purchase_Event_Count_YoY': 'New Properties_YoY',
        'FF_BRSubmit_Event_Count_YoY': 'Booking Requests_YoY',
        'FF_DMSubmit_Event_Count_YoY': 'Direct Messages_YoY',
        'FF_PhoneGet_Event_Count_YoY': 'Phone Reveals_YoY',
        'Housing Requests_YoY': 'Housing Requests_YoY',
        'Total Traveler Actions_YoY': 'Total Traveler Actons_YoY',
        'Traveler Value_YoY': 'Traveler Value_YoY',
        'Landlord Value_YoY': 'Landlord Value_YoY',
        'Sessions_YoY': 'Sessions_YoY',
        'Impressions_YoY': 'Impressions_YoY',
        'Clicks_YoY': 'Clicks_YoY',
        'Cost_YoY': 'Spend_YoY',
        'Lead_Conversion_YoY': 'Lead Conversion_YoY',
        'CPL_YoY': 'CPL_YoY',
        'CAC_YoY': 'CAC_YoY',
        'Traveler_Conversion_YoY': 'Traveler Conversion_YoY',
        'CPTA_YoY': 'CPTA_YoY',
        'ROAS_YoY': 'ROAS_YoY'
    }

    period_df.rename(columns=rename_map, inplace=True)

    # Define the desired column order
    columns_order = [
        'Spend_Actual', 'Spend_LW', 'Spend_WoW', 'Spend_YoY',
        'Leads_Actual', 'Leads_LW', 'Leads_WoW', 'Leads_YoY',
        'CPL_Actual', 'CPL_LW', 'CPL_WoW', 'CPL_YoY',
        'New Properties_Actual', 'New Properties_LW', 'New Properties_WoW', 'New Properties_YoY',
        'Lead Conversion_Actual', 'Lead Conversion_LW', 'Lead Conversion_WoW', 'Lead Conversion_YoY',
        'CAC_Actual', 'CAC_LW', 'CAC_WoW', 'CAC_YoY',
        'Booking Requests_Actual', 'Booking Requests_LW', 'Booking Requests_WoW', 'Booking Requests_YoY',
        'Direct Messages_Actual', 'Direct Messages_LW', 'Direct Messages_WoW', 'Direct Messages_YoY',
        'Phone Reveals_Actual', 'Phone Reveals_LW', 'Phone Reveals_WoW', 'Phone Reveals_YoY',
        'Housing Requests_Actual', 'Housing Requests_LW', 'Housing Requests_WoW', 'Housing Requests_YoY',
        'Traveler Conversion_Actual', 'Traveler Conversion_LW', 'Traveler Conversion_WoW', 'Traveler Conversion_YoY',
        'CPTA_Actual', 'CPTA_LW', 'CPTA_WoW', 'CPTA_YoY',
        'Traveler Value_Actual', 'Traveler Value_LW', 'Traveler Value_WoW', 'Traveler Value_YoY',
        'Landlord Value_Actual', 'Landlord Value_LW', 'Landlord Value_WoW', 'Landlord Value_YoY',
        'ROAS_Actual', 'ROAS_LW', 'ROAS_WoW', 'ROAS_YoY',
        'Impressions_Actual', 'Impressions_LW', 'Impressions_WoW', 'Impressions_YoY',
        'Clicks_Actual', 'Clicks_LW', 'Clicks_WoW', 'Clicks_YoY',
        'Sessions_Actual', 'Sessions_LW', 'Sessions_WoW', 'Sessions_YoY'
    ]

    # Reorder the columns according to the specified order
    period_df = period_df[['Channel_Group', 'Campaign_Group'] + [col for col in columns_order if col in period_df.columns]]

    # Add a new worksheet for the period
    ws = wb.create_sheet(title=period)

    # Set the title and sub-title
    ws['A1'] = f"GROWTH - PERFORMANCE REPORT - {period}"
    ws['A1'].font = Font(name="Aptos Display", size=18, bold=True, color="FFFFFF")
    ws['A1'].alignment = Alignment(horizontal="left")

    ws['A2'] = f"{start_date} to {end_date}"
    ws['A2'].font = Font(name="Aptos Display", size=16, bold=True, color="FFFFFF")
    ws['A2'].alignment = Alignment(horizontal="left")

    # Insert empty rows for padding between header and table
    ws.append([])  # Insert an empty row
    ws.append([])  # Insert another empty row

    # Insert the DataFrame into the worksheet
    for r in dataframe_to_rows(period_df, index=False, header=True):
        ws.append(r)
    
    # ------------- Formats of metrics based on column names --------------- #
    
    # Apply formats based on the column names
    for col in ws.iter_cols(min_col=3, max_col=ws.max_column, min_row=5, max_row=ws.max_row):
        header = col[0].value
        if header is not None:  # Ensure header is not None before processing
            print(f"Formatting column: {header}")  # Debugging line to check header
            for cell in col[1:]:
                if header.endswith('Actual') or header.endswith('LW'):
                    if 'Spend' in header or 'Traveler Value' in header or 'Landlord Value' in header:
                        cell.style = dollar_no_decimal
                    elif 'CPL' in header or 'CAC' in header:
                        cell.style = dollar_no_decimal
                    elif 'CPTA' in header:
                        cell.style = dollar_two_decimal
                    elif 'ROAS' in header:
                        cell.style = percent_no_decimal
                    elif 'Lead Conversion' in header or 'Traveler Conversion' in header:
                        cell.style = percent_no_decimal
                    else:
                        cell.style = comma_no_decimal
                elif header.endswith('WoW') or header.endswith('YoY'):
                    cell.style = percent_no_decimal
        
    # Insert the new row and split metric names
    # Now proceed to split the metric names into two rows
    # Create a list of the columns to format
    metric_names = [col.split("_")[0] for col in columns_order]
    unique_metric_names = sorted(set(metric_names), key=metric_names.index)

    # Insert a new row above the current column headers for the metric names
    ws.insert_rows(5)

    # Adjusted column offset to 2 to account for Channel_Group and Campaign_Group columns
    col_offset = 2  # Offset to account for Channel_Group and Campaign_Group columns
    for idx, metric_name in enumerate(unique_metric_names):
        start_col = idx * 4 + col_offset + 1
        end_col = start_col + 3

        # Set the metric name in the new row
        ws.cell(row=5, column=start_col, value=metric_name)

        # Merge and center the metric name cells over the four periods
        ws.merge_cells(start_row=5, start_column=start_col, end_row=5, end_column=end_col)
        ws.cell(row=5, column=start_col).alignment = Alignment(horizontal="center")

    # Adjust the original header row to show only the period names
    for col in ws.iter_cols(min_col=col_offset + 1, max_col=ws.max_column, min_row=6, max_row=6):
        period_label = col[0].value
        if period_label is not None:  # Ensure period_label is not None before processing
            col[0].value = period_label.split("_")[-1]
    
    # Define the border style for the outside border
    thin_side = Side(style='thin')

    # Apply outside borders around each metric column group
    for idx, metric_name in enumerate(unique_metric_names):
        start_col = idx * 4 + col_offset + 1
        end_col = start_col + 3
        
        # Apply borders from the top row (5) down to the last row in the data
        for row in range(5, ws.max_row + 1):
            # Set left and right borders for each cell in the group
            ws.cell(row=row, column=start_col).border = Border(left=thin_side)
            ws.cell(row=row, column=end_col).border = Border(right=thin_side)
            ws.cell(row=row, column=start_col).font = base_font  # Ensure font consistency
            ws.cell(row=row, column=end_col).font = base_font
        
        # Apply top and bottom borders and ensure side borders are maintained
        for col in range(start_col, end_col + 1):
            ws.cell(row=5, column=col).border = Border(top=thin_side, left=thin_side if col == start_col else None, right=thin_side if col == end_col else None)
            ws.cell(row=ws.max_row, column=col).border = Border(bottom=thin_side, left=thin_side if col == start_col else None, right=thin_side if col == end_col else None)
            ws.cell(row=5, column=col).font = base_font  # Ensure font consistency
            ws.cell(row=ws.max_row, column=col).font = base_font
            
    bold_font = Font(bold=True, name="Aptos Display", size=12)
    
    fill_color_1 = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    fill_color_2 = PatternFill(start_color="CAEDFB", end_color="CAEDFB", fill_type="solid")
    bold_font = Font(bold=True, name="Aptos Display", size=12)

    # Iterate through the metric names and apply the styles
    for idx, metric_name in enumerate(unique_metric_names):
        start_col = idx * 4 + col_offset + 1
        end_col = start_col + 3
        
        # Alternate between the two fill colors
        if idx % 2 == 0:
            fill_color = fill_color_1
        else:
            fill_color = fill_color_2
        
        # Apply the styles to the merged cells
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=5, column=col)
            cell.font = bold_font
            cell.fill = fill_color


    for row in range(7, ws.max_row + 1):  # Start from row 7 where data begins
        channel_group_cell = ws.cell(row=row, column=1)  # "Channel Group" is in column A (1)
        campaign_group_cell = ws.cell(row=row, column=2)  # "Campaign Group" is in column B (2)
        
        # Check if "SEM" or "Paid" is in the Channel Group or Campaign Group
        if "SEM" in str(channel_group_cell.value) or "Paid" in str(channel_group_cell.value):
            channel_group_cell.font = bold_font
            
        if "SEM" in str(campaign_group_cell.value) or "Paid" in str(campaign_group_cell.value):
            campaign_group_cell.font = bold_font

    
    # ------------------- Worksheet Color and Font Formatting ------------------- #
            
    # Apply White Fill Color to the Entire Workbook
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row + 50, min_col=1, max_col=ws.max_column + 50):
        for cell in row:
            cell.fill = white_fill
    
    # Apply Fill Color to Headers and Change Font Color to White
    header_fill = PatternFill(start_color="024991", end_color="024991", fill_type="solid")
    header_font = Font(color="FFFFFF", name="Aptos Display", size=12)

    for col in ws.iter_cols(min_col=1, max_col=74, min_row=6, max_row=6):
        for cell in col:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
    
    # Apply Alternating Row Colors
    alt_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    
    for row in range(7, ws.max_row + 1):
        if (row - 7) % 2 == 0:
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    cell.fill = alt_fill
    
    # Adjust Column Width for Columns A & B
    for col in ['A', 'B']:
        max_length = 0
        column = ws[col]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[col].width = adjusted_width
    
    # Apply Font Styling and Fill Color to the Title in the First Row
    ws['A1'].font = Font(name="Aptos Display", size=18, bold=False, color="FFFFFF")
    ws['A1'].alignment = Alignment(horizontal="left")
    for cell in ws["1:1"]:
        cell.fill = PatternFill(start_color="024991", end_color="024991", fill_type="solid")
    
    # Apply Font Styling and Fill Color to the Subtitle in the Second Row
    ws['A2'].font = Font(name="Aptos Display", size=16, bold=False, color="FFFFFF")
    ws['A2'].alignment = Alignment(horizontal="left")
    for cell in ws["2:2"]:
        cell.fill = PatternFill(start_color="D65488", end_color="D65488", fill_type="solid")

    # ---------------------- Add Formatting Logic Here ---------------------- #
    
    # ------------------------------------------------------------------------ #

    # Adjust column widths based on content
    for col in ws.iter_cols(min_col=1, max_col=ws.max_column, min_row=1, max_row=ws.max_row):
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
        
    # Adjust the width of each metric column
    min_width = 10  # Minimum width for the columns

    for col in ws.iter_cols(min_col=3, max_col=ws.max_column, min_row=7, max_row=ws.max_row):
        max_length = 0
        column = col[0].column_letter  # Get the column letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max((max_length + 2) * 1.2, min_width)  # Adjust width with padding
        ws.column_dimensions[column].width = adjusted_width
        
    # Freeze columns A & B so that they stay visible while scrolling to the right
    ws.freeze_panes = 'C7'


# Remove the default "Sheet" created by openpyxl
del wb['Sheet']

# Save the workbook
wb.save(excel_file_path)

print("Formatted Excel file with multiple sheets has been created and saved.")


# -------- Excel Workbook Creation and Formatting -------- #

# excel_file_path = 'ga4-reports/output/ga4_summary_metrics_formatted.xlsx' #/mnt/data/ga4_summary_metrics_formatted.xlsx'

# Load the existing Excel workbook
workbook = load_workbook(excel_file_path)

# Select the "EOW" worksheet
worksheet = workbook['EOW']

# Retrieve the start date from the second row, column A, in the "EOW" worksheet
eow_date_range = worksheet['A2'].value
eow_start_date_str = eow_date_range.split(' to ')[0]
eow_start_date = datetime.strptime(eow_start_date_str, '%m/%d/%Y')

# Calculate the ISO year and ISO week from the start date
iso_year = eow_start_date.isocalendar()[0]
iso_week = eow_start_date.isocalendar()[1]

# Filter the Google Sheets data based on the calculated ISO year and ISO week
filtered_data = [row for row in data if int(row[headers.index('Year')]) == iso_year and int(row[headers.index('ISO Week')]) == iso_week]

# Remove the "Year" and "ISO Week" columns from the filtered data
filtered_data = [[value for i, value in enumerate(row) if headers[i] not in ('Year', 'ISO Week')] for row in filtered_data]
filtered_headers = [header for header in headers if header not in ('Year', 'ISO Week')]

# Define the border style for the cells
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

# Specify the row and column where you want to start inserting the data
start_row = 46
start_column = 1  # Column A

# Insert the filtered headers into the Excel sheet and apply formatting
for col_num, header in enumerate(filtered_headers, start=start_column):
    cell = worksheet.cell(row=start_row, column=col_num, value=header)
    # Apply color fill, white font, and center alignment to the header row
    cell.fill = PatternFill(start_color="024991", end_color="024991", fill_type="solid")
    cell.font = Font(color="FFFFFF", name="Aptos Display", size=12)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# Merge columns C to R in the header row and apply formatting
worksheet.merge_cells(start_row=start_row, start_column=3, end_row=start_row, end_column=18)
header_cell = worksheet.cell(row=start_row, column=3)
header_cell.fill = PatternFill(start_color="024991", end_color="024991", fill_type="solid")
header_cell.font = Font(color="FFFFFF", name="Aptos Display", size=12)
header_cell.alignment = Alignment(horizontal='center', vertical='center')
header_cell.border = thin_border

# Insert the filtered data into the Excel sheet and apply formatting
for row_num, row_data in enumerate(filtered_data, start=start_row + 1):
    for col_num, value in enumerate(row_data, start=start_column):
        cell = worksheet.cell(row=row_num, column=col_num, value=value)
        cell.border = thin_border  # Apply borders to all data cells
        
        # Merge columns C to R for each row and left align
        if col_num == 3:  # Column C is index 3 when start_column is 1
            worksheet.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=18)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)  # Apply "Wrap Text"

        # Find "." in each cell in the commentary column starting from column C
        if col_num == 3 and isinstance(value, str):  # Commentary typically starts in column C
            cell.value = value.replace('.', '.\n')  # Replace "." with ".\n" to create a new line

    # Adjust row height to ensure all text is visible
    worksheet.row_dimensions[row_num].height = 80  # Increased the height more significantly

# Save the final Excel workbook with Google Sheets commentary included and formatted
workbook.save(excel_file_path)
print(f"Formatted and filtered data from Google Sheets (Year: {iso_year}, ISO Week: {iso_week}) has been written to {excel_file_path}, starting at row {start_row}, column {start_column} in the 'EOW' worksheet.")