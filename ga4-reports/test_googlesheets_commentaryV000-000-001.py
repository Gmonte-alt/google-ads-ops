# File name: ga4-reports/test_googlesheets_commentary.py
# version: V000-000-001
# output:
# Notes: Testing the set-up for the growth marketing cross channel reporting - now retrieves the data and inserts into an excel workbook first row of google sheets

import gspread
from google.oauth2.service_account import Credentials
from openpyxl import load_workbook

# Path to your service account JSON key file
SERVICE_ACCOUNT_FILE = 'authentication/green-carrier-423923-j0-ebd01cd30b56.json'

# Define the scopes required for the Google Sheets API
SCOPES = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive']

# Create a credentials object using the service account file
creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)

# Authenticate and initialize the gspread client
client = gspread.authorize(creds)

# Open the Google Sheet by its name or URL
sheet = client.open('FF - Growth - Weekly Commentary').sheet1

# Example: Read the values from the first row (assuming this is the data you want to add to Excel)
all_rows = sheet.get_all_values()
print("Row values from Google Sheets:", all_rows)

# Load the existing Excel workbook
excel_file_path = 'ga4-reports/output/ga4_summary_metrics_formatted.xlsx' # 'path_to_your_excel_file.xlsx'  # Replace with the actual path to your Excel file
workbook = load_workbook(excel_file_path)

# Select the worksheet you want to update
worksheet = workbook['EOW']  # Replace 'Sheet1' with your actual worksheet name

# Specify the row and column where you want to start inserting the data
start_row = 46
start_column = 1  # Column A

# Write all the data from Google Sheets into the Excel file, starting at the specified row and column
for row_num, row_data in enumerate(all_rows, start=start_row):
    for col_num, value in enumerate(row_data, start=start_column):
        worksheet.cell(row=row_num, column=col_num, value=value)

# Save the changes to the Excel file
workbook.save(excel_file_path)
print(f"Data from Google Sheets has been written to {excel_file_path}, starting at row {start_row}, column {start_column}.")