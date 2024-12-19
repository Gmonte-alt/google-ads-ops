# file name: ga4-reports/excel_workbook_summary_ga4_ad-platform_wtd.py
# version: V000-000-008
# output:
# Notes: First version of having four worksheets for WTD, EOW, MTD and QTD

import pandas as pd

# Load the CSV into a DataFrame
df = pd.read_csv('ga4-reports/output/ga4_summary_metrics_multiple_periods.csv')

# Preview the data to confirm it loaded correctly
print(df.head())


# Rename columns with '_Actual' Suffix

# Columns that need the '_Actual' suffix
columns_to_add_actual_suffix = [
    'FF_Lead_Event_Count', 'FF_Purchase_Event_Count', 'FF_BRSubmit_Event_Count',
    'FF_DMSubmit_Event_Count', 'FF_PhoneGet_Event_Count', 'Housing Requests',
    'Total Traveler Actions', 'Traveler Value', 'Landlord Value', 'Sessions',
    'Impressions', 'Clicks', 'Cost', 'Lead_Conversion', 'CPL', 'CAC',
    'Traveler_Conversion', 'CPTA', 'ROAS'
]

# Append '_Actual' suffix to the specified columns
df.rename(columns={col: f'{col}_Actual' for col in columns_to_add_actual_suffix}, inplace=True)

# Filter Data by Period and Create Worksheets

import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment

# List of periods
periods = ['WTD', 'EOW', 'MTD', 'QTD']

# Create a new Excel workbook
wb = openpyxl.Workbook()

# Iterate over each period
for period in periods:
    # Filter the DataFrame for the current period
    period_df = df[df['Period'] == period]

    # Determine the start and end dates
    start_date = period_df['Start_Date_Period'].min()
    end_date = period_df['End_Date_Period'].max()

    # Drop the unnecessary columns
    period_df = period_df.drop(columns=['Start_Date_Period', 'End_Date_Period', 'Period'])

    # Add a new worksheet for the period
    ws = wb.create_sheet(title=period)

    # Set the title and sub-title
    ws['A1'] = f"GROWTH - PERFORMANCE REPORT - {period}"
    ws['A1'].font = Font(name="Aptos Display", size=18, bold=True, color="FFFFFF")
    ws['A1'].alignment = Alignment(horizontal="left")

    ws['A2'] = f"{start_date} to {end_date}"
    ws['A2'].font = Font(name="Aptos Display", size=16, bold=True, color="FFFFFF")
    ws['A2'].alignment = Alignment(horizontal="left")

    # Insert the DataFrame into the worksheet
    for r in dataframe_to_rows(period_df, index=False, header=True):
        ws.append(r)


# Format the Worksheets

from openpyxl.styles import NamedStyle, PatternFill

# Apply formatting logic
for ws in wb.worksheets:
    # Example: Setting column widths
    for col in ws.iter_cols(min_col=1, max_col=ws.max_column, min_row=1, max_row=ws.max_row):
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    # Set the base font style and size across the entire worksheet
    base_font = Font(name="Aptos Display", size=12)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = base_font

    # Example: Applying alternating row colors
    alt_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

    for row in range(7, ws.max_row + 1):
        if (row - 7) % 2 == 0:
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    cell.fill = alt_fill

# Remove the default "Sheet" created by openpyxl
del wb['Sheet']

# Save the workbook
wb.save('ga4-reports/output/ga4_summary_metrics_formatted.xlsx')

print("Formatted Excel file with multiple sheets has been created and saved.")
