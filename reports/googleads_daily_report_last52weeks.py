# file name: reports/googleads_daily_report_last52weeks.py
# version: V000-000-016
# execution method: python googleads_daily_report_last52weeks.py
# Note: Builds on V000-000-015 to include event_value for ga4 data and visualizes Total Revenue & TD Revenue
#       switching metrics.all_conversions with metrics.all_conversions_by_conversion_date

import pandas as pd
import os
from google.ads.googleads.client import GoogleAdsClient
from google.ads.googleads.errors import GoogleAdsException
from datetime import datetime, timedelta
import re
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
import numpy as np

# Set the path to the Google Ads API configuration file
os.environ["GOOGLE_ADS_CONFIGURATION_FILE_PATH"] = "authentication/google-ads.yaml"

# Initialize the Google Ads client
client = GoogleAdsClient.load_from_storage()

def get_campaign_data(client, customer_id):
    ga_service = client.get_service("GoogleAdsService", version="v16")
    
    # Calculate date range for the past 52 weeks
    end_date = datetime.today().strftime('%Y-%m-%d')
    start_date = (datetime.today() - timedelta(weeks=60)).strftime('%Y-%m-%d')
    
    query = f"""
        SELECT
            segments.date,
            campaign.id,
            campaign.name,
            segments.device,
            segments.ad_network_type,
            segments.conversion_action,
            metrics.all_conversions_by_conversion_date
        FROM
            campaign
        WHERE
            segments.date BETWEEN '{start_date}' AND '{end_date}'
    """
    
    campaign_data = []
    try:
        response = ga_service.search_stream(customer_id=customer_id, query=query)
        for batch in response:
            for row in batch.results:
                campaign_data.append({
                    "Date": row.segments.date,
                    "Campaign ID": row.campaign.id,
                    "Campaign Name": row.campaign.name,
                    "Device": row.segments.device.name,
                    "Network": row.segments.ad_network_type.name,
                    "Conversion Action ID": row.segments.conversion_action,
                    "Conversions": row.metrics.all_conversions_by_conversion_date
                })
    except GoogleAdsException as ex:
        print(f"Request failed with status {ex.error.code().name} and includes the following errors:")
        for error in ex.failure.errors:
            print(f"\tError with message {error.message}.")
            if error.location:
                for field_path_element in error.location.field_path_elements:
                    print(f"\t\tOn field: {field_path_element.field_name}")
        return None

    # Convert dataset to DataFrame
    campaign_df = pd.DataFrame(campaign_data)
    
    return campaign_df

def get_conversion_action_data(client, customer_id):
    ga_service = client.get_service("GoogleAdsService", version="v16")
    
    query = """
        SELECT
            conversion_action.id,
            conversion_action.name
        FROM
            conversion_action
    """
    
    conversion_action_data = []
    try:
        response = ga_service.search_stream(customer_id=customer_id, query=query)
        for batch in response:
            for row in batch.results:
                conversion_action_data.append({
                    "Conversion Action ID": row.conversion_action.id,
                    "Conversion Action Name": row.conversion_action.name
                })
    except GoogleAdsException as ex:
        print(f"Request failed with status {ex.error.code().name} and includes the following errors:")
        for error in ex.failure.errors:
            print(f"\tError with message {error.message}.")
            if error.location:
                for field_path_element in error.location.field_path_elements:
                    print(f"\t\tOn field: {field_path_element.field_name}")
        return None

    # Convert dataset to DataFrame
    conversion_action_df = pd.DataFrame(conversion_action_data)
    
    return conversion_action_df

def get_clicks_impressions_spend_data(client, customer_id):
    ga_service = client.get_service("GoogleAdsService", version="v16")
    
    # Calculate date range for the past 52 weeks
    end_date = datetime.today().strftime('%Y-%m-%d')
    start_date = (datetime.today() - timedelta(weeks=60)).strftime('%Y-%m-%d')
    
    query = f"""
        SELECT
            segments.date,
            campaign.id,
            campaign.name,
            segments.device,
            segments.ad_network_type,
            metrics.clicks,
            metrics.impressions,
            metrics.cost_micros,
            metrics.search_click_share,
            metrics.search_budget_lost_top_impression_share,
            metrics.search_budget_lost_impression_share,
            metrics.search_budget_lost_absolute_top_impression_share,
            metrics.search_absolute_top_impression_share,
            metrics.top_impression_percentage
        FROM
            campaign
        WHERE
            segments.date BETWEEN '{start_date}' AND '{end_date}'
    """
    
    clicks_impressions_spend_data = []
    try:
        response = ga_service.search_stream(customer_id=customer_id, query=query)
        for batch in response:
            for row in batch.results:
                clicks_impressions_spend_data.append({
                    "Date": row.segments.date,
                    "Campaign ID": row.campaign.id,
                    "Campaign Name": row.campaign.name,
                    "Device": row.segments.device.name,
                    "Network": row.segments.ad_network_type.name,
                    "Clicks": row.metrics.clicks,
                    "Impressions": row.metrics.impressions,
                    "Cost": row.metrics.cost_micros / 1e6,  # Convert micros to currency
                    "absolute_top_impression_share": row.metrics.search_absolute_top_impression_share,
                    "search_click_share": row.metrics.search_click_share
                })
    except GoogleAdsException as ex:
        print(f"Request failed with status {ex.error.code().name} and includes the following errors:")
        for error in ex.failure.errors:
            print(f"\tError with message {error.message}.")
            if error.location:
                for field_path_element in error.location.field_path_elements:
                    print(f"\t\tOn field: {field_path_element.field_name}")
        return None

    # Convert dataset to DataFrame
    clicks_impressions_spend_df = pd.DataFrame(clicks_impressions_spend_data)
    
    return clicks_impressions_spend_df

def extract_conversion_action_id(conversion_action_str):
    match = re.search(r'conversionActions/(\d+)', conversion_action_str)
    return match.group(1) if match else None

def save_to_excel(dataframe, filename, sheet_name='Sheet1'):
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        dataframe.to_excel(writer, sheet_name=sheet_name, index=False)

def save_combined_to_excel(data_dict, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet"
    
    sheet_names = []
    for date, data in data_dict.items():
        sheet_names.append(date)
        
        # Create a new sheet for each date
        new_ws = wb.create_sheet(title=date)
        
        row_start = 1
        for title, data in data.items():
            # Add title
            new_ws.cell(row=row_start, column=1, value=title).font = Font(size=14, bold=True)
            
            # Convert DataFrame to rows and append to worksheet
            for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), start=row_start + 1):
                for c_idx, value in enumerate(row, start=1):
                    # Ensure the value is scalar
                    if isinstance(value, np.ndarray):
                        value = value.item()  # Convert numpy array to scalar
                    
                    new_ws.cell(row=r_idx, column=c_idx, value=value)
            
            # Move to next start row after the table, plus some space
            row_start = new_ws.max_row + 3
    
    # Populate the "Sheet" worksheet with the names of the other worksheets
    for idx, sheet_name in enumerate(sheet_names, start=2):  # Start from cell B2
        ws.cell(row=2, column=idx, value=sheet_name)
    
    wb.save(filename)


# Calculate ratio metrics
def safe_divide(numerator, denominator):
    with np.errstate(divide='ignore', invalid='ignore'):
        result = np.where(denominator != 0, numerator / denominator, 0)
    return result

# Assume final_data is already loaded into a pandas DataFrame
# final_data = pd.read_csv("path_to_final_data.csv")

def calculate_summary_metrics(final_data, date):
    # Define the metrics
    metrics = [
        "Furnished Finder - GA4 (web) FF Purchase",
        "Furnished Finder - GA4 (web) FF Lead",
        "Cost",
        "Clicks",
        "Furnished Finder - GA4 (web) FF-BRSubmit",
        "Furnished Finder - GA4 (web) FF-DMSubmit",
        "Furnished Finder - GA4 (web) FF-PhoneGet",
        "ga4_FF_Purchase_Event_Count",
        "ga4_FF_Lead_Event_Count",
        "ga4_FF_BRSubmit_Event_Count",
        "ga4_FF_DMSubmit_Event_Count",
        "ga4_FF_PhoneGet_Event_Count"
    ]

    # Ensure 'Date' is in datetime format first
    final_data.loc[:, 'Date'] = pd.to_datetime(final_data.loc[:, 'Date'], errors='coerce')
    
    # Ensure 'Date' is in string format
    final_data.loc[:, 'Date'] = final_data.loc[:, 'Date'].dt.strftime('%Y-%m-%d')
    
    # Calculate date references
    today = datetime.today()
    yesterday = date
    yesterday_str = yesterday.strftime('%Y-%m-%d')
    day_before_yesterday = yesterday - timedelta(days=1)
    day_before_yesterday_str = day_before_yesterday.strftime('%Y-%m-%d')
    last_week_same_day = yesterday - timedelta(days=7)
    last_week_same_day_str = last_week_same_day.strftime('%Y-%m-%d')
    two_weeks_ago_same_day = yesterday - timedelta(days=14)
    two_weeks_ago_same_day_str = two_weeks_ago_same_day.strftime('%Y-%m-%d')
    
    # Calculate the date 26 weeks (182 days) ago from yesterday
    date_26_weeks_ago = yesterday - timedelta(weeks=26)
    date_26_weeks_ago_str = date_26_weeks_ago.strftime('%Y-%m-%d')
    
    # Calculate the date 52 weeks (364 days) ago from yesterday
    date_52_weeks_ago = yesterday - timedelta(weeks=52)
    date_52_weeks_ago_str = date_52_weeks_ago.strftime('%Y-%m-%d')

    # Get the ISO calendar week and weekday for yesterday
    iso_year, iso_week, iso_weekday = yesterday.isocalendar()

    # Calculate the same day of the same ISO week from the previous year
    last_year_same_iso_week = datetime.strptime(f'{iso_year-1}-W{iso_week}-{iso_weekday}', "%G-W%V-%u")
    last_year_same_iso_week_str = last_year_same_iso_week.strftime('%Y-%m-%d')
    
    # Function to calculate the start of last month
    def calculate_start_of_last_month(date=None):
        if date is None:
            date = datetime.today()
        first_day_of_this_month = date.replace(day=1)
        last_day_of_last_month = first_day_of_this_month - timedelta(days=1)
        start_of_last_month = last_day_of_last_month.replace(day=1)
        return start_of_last_month
    
    # Function to calculate the start of the month prior to last month
    def calculate_start_of_month_before_last(date=None):
        if date is None:
            date = datetime.today()
        start_of_last_month_date = calculate_start_of_last_month(date)
        last_day_of_month_before_last = start_of_last_month_date - timedelta(days=1)
        start_of_month_before_last = last_day_of_month_before_last.replace(day=1)
        return start_of_month_before_last
    
    # Calculate start of last month
    start_of_last_month_date = calculate_start_of_last_month(yesterday)
    start_of_last_month_str = start_of_last_month_date.strftime('%Y-%m-%d')
    
    # Calculate the equivalent day of the month last month
    def safe_replace_day(date, day):
        try:
            return date.replace(day=day)
        except ValueError:
            # This handles the case where the day is out of range for the month
            return date.replace(day=1) + timedelta(days=day - 1)
    
    # Calculate the equivalent day of the month last month
    last_month_same_day = safe_replace_day(start_of_last_month_date, yesterday.day)
    last_month_same_day_str = last_month_same_day.strftime('%Y-%m-%d')
    
    # Calculate start of the month prior to last month
    start_of_month_before_last_date = calculate_start_of_month_before_last(yesterday)
    start_of_month_before_last_str = start_of_month_before_last_date.strftime('%Y-%m-%d')

    # Calculate equivalent day of the month before last month
    month_before_last_same_day = safe_replace_day(start_of_month_before_last_date, yesterday.day)
    month_before_last_same_day_str = month_before_last_same_day.strftime('%Y-%m-%d')

    # Calculate start of the current week
    start_of_week = yesterday - timedelta(days=iso_weekday - 1)
    start_of_week_str = start_of_week.strftime('%Y-%m-%d')

    # Calculate start of the previous week
    start_of_last_week = start_of_week - timedelta(weeks=1)
    start_of_last_week_str = start_of_last_week.strftime('%Y-%m-%d')
    
    # Calculate start of the week two weeks ago
    start_of_two_weeks_ago = start_of_week - timedelta(weeks=2)
    start_of_two_weeks_ago_str = start_of_two_weeks_ago.strftime('%Y-%m-%d')
    
    # Calculate the start of the week 26 weeks ago
    start_of_26_weeks_ago = start_of_week - timedelta(weeks=26)
    start_of_26_weeks_ago_str = start_of_26_weeks_ago.strftime('%Y-%m-%d')
    
    # Calculate the start of the week 52 weeks ago
    start_of_52_weeks_ago = start_of_week - timedelta(weeks=52)
    start_of_52_weeks_ago_str = start_of_52_weeks_ago.strftime('%Y-%m-%d')

    # Calculate the start of the same ISO week from the previous year
    start_of_last_year_iso_week = datetime.strptime(f'{iso_year-1}-W{iso_week}-1', "%G-W%V-%u")
    #start_of_last_year_iso_week_end = start_of_last_year_iso_week + timedelta(weeks=1)

    # Calculate start of the current month
    start_of_month = yesterday.replace(day=1)
    start_of_month_str = start_of_month.strftime('%Y-%m-%d')

    # Calculate the same month from the previous year
    start_of_last_year_month = yesterday.replace(year=yesterday.year - 1, month=yesterday.month, day=1)
    start_of_last_year_month_str = start_of_last_year_month.strftime('%Y-%m-%d')
    # Calculate equivalent day of the month from the previous year
    last_year_month_same_day = safe_replace_day(start_of_last_year_month, yesterday.day)
    last_year_month_same_day_str = last_year_month_same_day.strftime('%Y-%m-%d')
    ## end_of_last_year_month = (start_of_last_year_month.replace(day=1) + timedelta(days=31)).replace(day=1)
    ## end_of_last_year_month_str = end_of_last_year_month.strftime('%Y-%m-%d')

    # Filter data for the relevant days into Pandas Series
    yesterday_data = final_data[final_data['Date'] == yesterday_str]
    day_before_data = final_data[final_data['Date'] == day_before_yesterday_str]
    last_week_data = final_data[final_data['Date'] == last_week_same_day_str]
    two_weeks_ago_data = final_data[final_data['Date'] == two_weeks_ago_same_day_str]
    last_year_same_iso_week_data = final_data[final_data['Date'] == last_year_same_iso_week_str]

    # Filter data for the last 26 weeks
    last_26_weeks_data = final_data[(final_data['Date'] >= date_26_weeks_ago_str) & (final_data['Date'] <= yesterday_str)]
    # Further filter data for the same weekday over the last 26 weeks
    same_weekday_last_26_weeks = last_26_weeks_data[last_26_weeks_data['Date'].apply(lambda x: x.weekday() == yesterday.weekday())]

    # Filter data for the last 52 weeks
    last_52_weeks_data = final_data[(final_data['Date'] >= date_52_weeks_ago_str) & (final_data['Date'] <= yesterday_str)]
    # Further filter data for the same weekday over the last 52 weeks
    same_weekday_last_52_weeks = last_52_weeks_data[last_52_weeks_data['Date'].apply(lambda x: x.weekday() == yesterday.weekday())]

    # Filter data for all time in the dataframe
    same_weekday_all_time_in_dataframe = final_data[final_data['Date'].apply(lambda x: x.weekday() == yesterday.weekday())]

    # Filter data for week-to-date (WTD) metrics
    wtd_data = final_data[(final_data['Date'] >= start_of_week_str) & (final_data['Date'] <= yesterday_str)]
    last_week_wtd_data = final_data[(final_data['Date'] >= start_of_last_week_str) & (final_data['Date'] <= last_week_same_day_str)]
    two_weeks_ago_wtd_data = final_data[(final_data['Date'] >= start_of_two_weeks_ago_str) & (final_data['Date'] <= two_weeks_ago_same_day_str)]

    # Filter data for the last 26 full weeks
    start_26_weeks_ago_data = final_data[(final_data['Date'] >= start_of_26_weeks_ago_str) & (final_data['Date'] <= yesterday_str)]
    # Filter data for the last 52 full weeks
    start_52_weeks_ago_data = final_data[(final_data['Date'] >= start_of_52_weeks_ago_str) & (final_data['Date'] <= yesterday_str)]

    last_year_same_iso_week_wtd_data = final_data[(final_data['Date'] >= start_of_last_year_iso_week.strftime('%Y-%m-%d')) & (final_data['Date'] <= last_year_same_iso_week_str)]

    # Filter data for month-to-date (MTD) metrics
    mtd_data = final_data[(final_data['Date'] >= start_of_month_str) & (final_data['Date'] <= yesterday_str)]
    last_month_mtd_data = final_data[(final_data['Date'] >= start_of_last_month_str) & (final_data['Date'] <= last_month_same_day_str)]
    two_months_ago_mtd_data = final_data[(final_data['Date'] >= start_of_month_before_last_str) & (final_data['Date'] <= month_before_last_same_day_str)]
    last_year_month_mtd_data = final_data[(final_data['Date'] >= start_of_last_year_month_str) & (final_data['Date'] <= last_year_month_same_day_str)]


    # Sum the metrics for the filtered data
    yesterday_summary = yesterday_data[metrics].sum().fillna(0)
    day_before_summary = day_before_data[metrics].sum().fillna(0)
    last_week_summary = last_week_data[metrics].sum().fillna(0)
    two_weeks_ago_summary = two_weeks_ago_data[metrics].sum().fillna(0)
    last_year_same_iso_week_summary = last_year_same_iso_week_data[metrics].sum().fillna(0)
    avg_same_weekday_26wk_summary = same_weekday_last_26_weeks.groupby('Date')[metrics].sum().mean().fillna(0)
    avg_same_weekday_52wk_summary = same_weekday_last_52_weeks.groupby('Date')[metrics].sum().mean().fillna(0)

    # Sum the metrics for WTD data
    wtd_summary = wtd_data[metrics].sum().fillna(0)
    last_week_wtd_summary = last_week_wtd_data[metrics].sum().fillna(0)
    two_weeks_ago_wtd_summary = two_weeks_ago_wtd_data[metrics].sum().fillna(0)
    last_year_same_iso_week_wtd_summary = last_year_same_iso_week_wtd_data[metrics].sum().fillna(0)
    avg_same_wtd_26wk_summary = start_26_weeks_ago_data.groupby('Date')[metrics].sum().mean().fillna(0)
    avg_same_wtd_52wk_summary = start_52_weeks_ago_data.groupby('Date')[metrics].sum().mean().fillna(0)

    # Sum the metrics for MTD data
    mtd_summary = mtd_data[metrics].sum().fillna(0)
    last_month_mtd_summary = last_month_mtd_data[metrics].sum().fillna(0)
    two_months_ago_mtd_summary = two_months_ago_mtd_data[metrics].sum().fillna(0)
    last_year_month_mtd_summary = last_year_month_mtd_data[metrics].sum().fillna(0)


    # Ratio metrics for the filtered data yesterday_summary
    cac = safe_divide(yesterday_summary['Cost'], yesterday_summary["Furnished Finder - GA4 (web) FF Purchase"])
    cpl = safe_divide(yesterday_summary['Cost'], yesterday_summary["Furnished Finder - GA4 (web) FF Lead"])
    cpc = safe_divide(yesterday_summary['Cost'], yesterday_summary['Clicks'])
    cost_per_brsubmit = safe_divide(yesterday_summary['Cost'], yesterday_summary["Furnished Finder - GA4 (web) FF-BRSubmit"])
    brsubmit_clicks_cvr = safe_divide(yesterday_summary["Furnished Finder - GA4 (web) FF-BRSubmit"], yesterday_summary['Clicks'])
    lead_cvr = safe_divide(yesterday_summary["Furnished Finder - GA4 (web) FF Purchase"], yesterday_summary["Furnished Finder - GA4 (web) FF Lead"])
    click_cvr = safe_divide(yesterday_summary["Furnished Finder - GA4 (web) FF Lead"], yesterday_summary['Clicks'])

    ga4_cac = safe_divide(yesterday_summary['Cost'], yesterday_summary["ga4_FF_Purchase_Event_Count"])
    ga4_lead_cvr = safe_divide(yesterday_summary["ga4_FF_Purchase_Event_Count"], yesterday_summary["ga4_FF_Lead_Event_Count"])

    # Ratio metrics for the filtered data day_before_summary
    day_before_cac = safe_divide(day_before_summary['Cost'], day_before_summary["Furnished Finder - GA4 (web) FF Purchase"])
    day_before_cpl = safe_divide(day_before_summary['Cost'], day_before_summary["Furnished Finder - GA4 (web) FF Lead"])
    day_before_cpc = safe_divide(day_before_summary['Cost'], day_before_summary['Clicks'])
    day_before_cost_per_brsubmit = safe_divide(day_before_summary['Cost'], day_before_summary["Furnished Finder - GA4 (web) FF-BRSubmit"])
    day_before_brsubmit_clicks_cvr = safe_divide(day_before_summary["Furnished Finder - GA4 (web) FF-BRSubmit"], day_before_summary['Clicks'])
    day_before_lead_cvr = safe_divide(day_before_summary["Furnished Finder - GA4 (web) FF Purchase"], day_before_summary["Furnished Finder - GA4 (web) FF Lead"])
    day_before_click_cvr = safe_divide(day_before_summary["Furnished Finder - GA4 (web) FF Lead"], day_before_summary['Clicks'])

    day_before_ga4_cac = safe_divide(day_before_summary['Cost'], day_before_summary["ga4_FF_Purchase_Event_Count"])
    day_before_ga4_lead_cvr = safe_divide(day_before_summary["ga4_FF_Purchase_Event_Count"], day_before_summary["ga4_FF_Lead_Event_Count"])

    # Ratio metrics for the filtered data last_week_summary
    last_week_cac = safe_divide(last_week_summary['Cost'], last_week_summary["Furnished Finder - GA4 (web) FF Purchase"])
    last_week_cpl = safe_divide(last_week_summary['Cost'], last_week_summary["Furnished Finder - GA4 (web) FF Lead"])
    last_week_cpc = safe_divide(last_week_summary['Cost'], last_week_summary['Clicks'])
    last_week_lead_cvr = safe_divide(last_week_summary["Furnished Finder - GA4 (web) FF Purchase"], last_week_summary["Furnished Finder - GA4 (web) FF Lead"])
    last_week_click_cvr = safe_divide(last_week_summary["Furnished Finder - GA4 (web) FF Lead"], last_week_summary['Clicks'])
    last_week_cost_per_brsubmit = safe_divide(last_week_summary['Cost'], last_week_summary["Furnished Finder - GA4 (web) FF-BRSubmit"])
    last_week_brsubmit_clicks_cvr = safe_divide(last_week_summary["Furnished Finder - GA4 (web) FF-BRSubmit"], last_week_summary['Clicks'])

    last_week_ga4_cac = safe_divide(last_week_summary['Cost'], last_week_summary["ga4_FF_Purchase_Event_Count"])
    last_week_ga4_lead_cvr = safe_divide(last_week_summary["ga4_FF_Purchase_Event_Count"], last_week_summary["ga4_FF_Lead_Event_Count"])

    # Ratio metrics for the filtered data two_weeks_ago_summary
    two_weeks_ago_cac = safe_divide(two_weeks_ago_summary['Cost'], two_weeks_ago_summary["Furnished Finder - GA4 (web) FF Purchase"])
    two_weeks_ago_cpl = safe_divide(two_weeks_ago_summary['Cost'], two_weeks_ago_summary["Furnished Finder - GA4 (web) FF Lead"])
    two_weeks_ago_cpc = safe_divide(two_weeks_ago_summary['Cost'], two_weeks_ago_summary['Clicks'])
    two_weeks_ago_lead_cvr = safe_divide(two_weeks_ago_summary["Furnished Finder - GA4 (web) FF Purchase"], two_weeks_ago_summary["Furnished Finder - GA4 (web) FF Lead"])
    two_weeks_ago_click_cvr = safe_divide(two_weeks_ago_summary["Furnished Finder - GA4 (web) FF Lead"], two_weeks_ago_summary['Clicks'])
    two_weeks_ago_cost_per_brsubmit = safe_divide(two_weeks_ago_summary['Cost'], two_weeks_ago_summary["Furnished Finder - GA4 (web) FF-BRSubmit"])
    two_weeks_ago_brsubmit_clicks_cvr = safe_divide(two_weeks_ago_summary["Furnished Finder - GA4 (web) FF-BRSubmit"], two_weeks_ago_summary['Clicks'])

    two_weeks_ago_ga4_cac = safe_divide(two_weeks_ago_summary['Cost'], two_weeks_ago_summary["ga4_FF_Purchase_Event_Count"])
    two_weeks_ago_ga4_lead_cvr = safe_divide(two_weeks_ago_summary["ga4_FF_Purchase_Event_Count"], two_weeks_ago_summary["ga4_FF_Lead_Event_Count"])

    # Ratio metrics for the filtered data avg_same_weekday_summary change to avg_same_weekday_26wk_summary
    avg_same_weekday_26wk_cac = safe_divide(avg_same_weekday_26wk_summary['Cost'], avg_same_weekday_26wk_summary["Furnished Finder - GA4 (web) FF Purchase"])
    avg_same_weekday_26wk_cpl = safe_divide(avg_same_weekday_26wk_summary['Cost'], avg_same_weekday_26wk_summary["Furnished Finder - GA4 (web) FF Lead"])
    avg_same_weekday_26wk_cpc = safe_divide(avg_same_weekday_26wk_summary['Cost'], avg_same_weekday_26wk_summary['Clicks'])
    avg_same_weekday_26wk_lead_cvr = safe_divide(avg_same_weekday_26wk_summary["Furnished Finder - GA4 (web) FF Purchase"], avg_same_weekday_26wk_summary["Furnished Finder - GA4 (web) FF Lead"])
    avg_same_weekday_26wk_click_cvr = safe_divide(avg_same_weekday_26wk_summary["Furnished Finder - GA4 (web) FF Lead"], avg_same_weekday_26wk_summary['Clicks'])
    avg_same_weekday_26wk_cost_per_brsubmit = safe_divide(avg_same_weekday_26wk_summary['Cost'], avg_same_weekday_26wk_summary["Furnished Finder - GA4 (web) FF-BRSubmit"])
    avg_same_weekday_26wk_brsubmit_clicks_cvr = safe_divide(avg_same_weekday_26wk_summary["Furnished Finder - GA4 (web) FF-BRSubmit"], avg_same_weekday_26wk_summary['Clicks'])

    avg_same_weekday_26wk_ga4_cac = safe_divide(avg_same_weekday_26wk_summary['Cost'], avg_same_weekday_26wk_summary["ga4_FF_Purchase_Event_Count"])
    avg_same_weekday_26wk_ga4_lead_cvr = safe_divide(avg_same_weekday_26wk_summary["ga4_FF_Purchase_Event_Count"], avg_same_weekday_26wk_summary["ga4_FF_Lead_Event_Count"])

    # Ratio metrics for the filtered data avg_same_weekday_summary change to avg_same_weekday_52wk_summary
    avg_same_weekday_52wk_cac = safe_divide(avg_same_weekday_52wk_summary['Cost'], avg_same_weekday_52wk_summary["Furnished Finder - GA4 (web) FF Purchase"])
    avg_same_weekday_52wk_cpl = safe_divide(avg_same_weekday_52wk_summary['Cost'], avg_same_weekday_52wk_summary["Furnished Finder - GA4 (web) FF Lead"])
    avg_same_weekday_52wk_cpc = safe_divide(avg_same_weekday_52wk_summary['Cost'], avg_same_weekday_52wk_summary['Clicks'])
    avg_same_weekday_52wk_lead_cvr = safe_divide(avg_same_weekday_52wk_summary["Furnished Finder - GA4 (web) FF Purchase"], avg_same_weekday_52wk_summary["Furnished Finder - GA4 (web) FF Lead"])
    avg_same_weekday_52wk_click_cvr = safe_divide(avg_same_weekday_52wk_summary["Furnished Finder - GA4 (web) FF Lead"], avg_same_weekday_52wk_summary['Clicks'])
    avg_same_weekday_52wk_cost_per_brsubmit = safe_divide(avg_same_weekday_52wk_summary['Cost'], avg_same_weekday_52wk_summary["Furnished Finder - GA4 (web) FF-BRSubmit"])
    avg_same_weekday_52wk_brsubmit_clicks_cvr = safe_divide(avg_same_weekday_52wk_summary["Furnished Finder - GA4 (web) FF-BRSubmit"], avg_same_weekday_52wk_summary['Clicks'])

    avg_same_weekday_52wk_ga4_cac = safe_divide(avg_same_weekday_52wk_summary['Cost'], avg_same_weekday_52wk_summary["ga4_FF_Purchase_Event_Count"])
    avg_same_weekday_52wk_ga4_lead_cvr = safe_divide(avg_same_weekday_52wk_summary["ga4_FF_Purchase_Event_Count"], avg_same_weekday_52wk_summary["ga4_FF_Lead_Event_Count"])

    # Ratio metrics for the filtered data last_year_same_iso_week_summary
    last_year_same_iso_week_cac = safe_divide(last_year_same_iso_week_summary['Cost'], last_year_same_iso_week_summary["Furnished Finder - GA4 (web) FF Purchase"])
    last_year_same_iso_week_cpl = safe_divide(last_year_same_iso_week_summary['Cost'], last_year_same_iso_week_summary["Furnished Finder - GA4 (web) FF Lead"])
    last_year_same_iso_week_cpc = safe_divide(last_year_same_iso_week_summary['Cost'], last_year_same_iso_week_summary['Clicks'])
    last_year_same_iso_week_lead_cvr = safe_divide(last_year_same_iso_week_summary["Furnished Finder - GA4 (web) FF Purchase"], last_year_same_iso_week_summary["Furnished Finder - GA4 (web) FF Lead"])
    last_year_same_iso_week_click_cvr = safe_divide(last_year_same_iso_week_summary["Furnished Finder - GA4 (web) FF Lead"], last_year_same_iso_week_summary['Clicks'])
    last_year_same_iso_week_cost_per_brsubmit = safe_divide(last_year_same_iso_week_summary['Cost'], last_year_same_iso_week_summary["Furnished Finder - GA4 (web) FF-BRSubmit"])
    last_year_same_iso_week_brsubmit_clicks_cvr = safe_divide(last_year_same_iso_week_summary["Furnished Finder - GA4 (web) FF-BRSubmit"], last_year_same_iso_week_summary['Clicks'])

    last_year_same_iso_week_ga4_cac = safe_divide(last_year_same_iso_week_summary['Cost'], last_year_same_iso_week_summary["ga4_FF_Purchase_Event_Count"])
    last_year_same_iso_week_ga4_lead_cvr = safe_divide(last_year_same_iso_week_summary["ga4_FF_Purchase_Event_Count"], last_year_same_iso_week_summary["ga4_FF_Lead_Event_Count"])

    # Ratio metrics for the filtered data wtd_summary
    wtd_cac = safe_divide(wtd_summary['Cost'], wtd_summary["Furnished Finder - GA4 (web) FF Purchase"])
    wtd_cpl = safe_divide(wtd_summary['Cost'], wtd_summary["Furnished Finder - GA4 (web) FF Lead"])
    wtd_cpc = safe_divide(wtd_summary['Cost'], wtd_summary['Clicks'])
    wtd_lead_cvr = safe_divide(wtd_summary["Furnished Finder - GA4 (web) FF Purchase"], wtd_summary["Furnished Finder - GA4 (web) FF Lead"])
    wtd_click_cvr = safe_divide(wtd_summary["Furnished Finder - GA4 (web) FF Lead"], wtd_summary['Clicks'])
    wtd_cost_per_brsubmit = safe_divide(wtd_summary['Cost'], wtd_summary["Furnished Finder - GA4 (web) FF-BRSubmit"])
    wtd_brsubmit_clicks_cvr = safe_divide(wtd_summary["Furnished Finder - GA4 (web) FF-BRSubmit"], wtd_summary['Clicks'])

    wtd_ga4_cac = safe_divide(wtd_summary['Cost'], wtd_summary["ga4_FF_Purchase_Event_Count"])
    wtd_ga4_lead_cvr = safe_divide(wtd_summary["ga4_FF_Purchase_Event_Count"], wtd_summary["ga4_FF_Lead_Event_Count"])

    # Ratio metrics for the filtered data last_week_wtd_summary
    last_week_wtd_cac = safe_divide(last_week_wtd_summary['Cost'], last_week_wtd_summary["Furnished Finder - GA4 (web) FF Purchase"])
    last_week_wtd_cpl = safe_divide(last_week_wtd_summary['Cost'], last_week_wtd_summary["Furnished Finder - GA4 (web) FF Lead"])
    last_week_wtd_cpc = safe_divide(last_week_wtd_summary['Cost'], last_week_wtd_summary['Clicks'])
    last_week_wtd_lead_cvr = safe_divide(last_week_wtd_summary["Furnished Finder - GA4 (web) FF Purchase"], last_week_wtd_summary["Furnished Finder - GA4 (web) FF Lead"])
    last_week_wtd_click_cvr = safe_divide(last_week_wtd_summary["Furnished Finder - GA4 (web) FF Lead"], last_week_wtd_summary['Clicks'])
    last_week_wtd_cost_per_brsubmit = safe_divide(last_week_wtd_summary['Cost'], last_week_wtd_summary["Furnished Finder - GA4 (web) FF-BRSubmit"])
    last_week_wtd_brsubmit_clicks_cvr = safe_divide(last_week_wtd_summary["Furnished Finder - GA4 (web) FF-BRSubmit"], last_week_wtd_summary['Clicks'])

    last_week_wtd_ga4_cac = safe_divide(last_week_wtd_summary['Cost'], last_week_wtd_summary["ga4_FF_Purchase_Event_Count"])
    last_week_wtd_ga4_lead_cvr = safe_divide(last_week_wtd_summary["ga4_FF_Purchase_Event_Count"], last_week_wtd_summary["ga4_FF_Lead_Event_Count"])

    # Ratio metrics for the filtered data two_weeks_ago_wtd_summary
    two_weeks_ago_wtd_cac = safe_divide(two_weeks_ago_wtd_summary['Cost'], two_weeks_ago_wtd_summary["Furnished Finder - GA4 (web) FF Purchase"])
    two_weeks_ago_wtd_cpl = safe_divide(two_weeks_ago_wtd_summary['Cost'], two_weeks_ago_wtd_summary["Furnished Finder - GA4 (web) FF Lead"])
    two_weeks_ago_wtd_cpc = safe_divide(two_weeks_ago_wtd_summary['Cost'], two_weeks_ago_wtd_summary['Clicks'])
    two_weeks_ago_wtd_lead_cvr = safe_divide(two_weeks_ago_wtd_summary["Furnished Finder - GA4 (web) FF Purchase"], two_weeks_ago_wtd_summary["Furnished Finder - GA4 (web) FF Lead"])
    two_weeks_ago_wtd_click_cvr = safe_divide(two_weeks_ago_wtd_summary["Furnished Finder - GA4 (web) FF Lead"], two_weeks_ago_wtd_summary['Clicks'])
    two_weeks_ago_wtd_cost_per_brsubmit = safe_divide(two_weeks_ago_wtd_summary['Cost'], two_weeks_ago_wtd_summary["Furnished Finder - GA4 (web) FF-BRSubmit"])
    two_weeks_ago_wtd_brsubmit_clicks_cvr = safe_divide(two_weeks_ago_wtd_summary["Furnished Finder - GA4 (web) FF-BRSubmit"], two_weeks_ago_wtd_summary['Clicks'])

    two_weeks_ago_wtd_ga4_cac = safe_divide(two_weeks_ago_wtd_summary['Cost'], two_weeks_ago_wtd_summary["ga4_FF_Purchase_Event_Count"])
    two_weeks_ago_wtd_ga4_lead_cvr = safe_divide(two_weeks_ago_wtd_summary["ga4_FF_Purchase_Event_Count"], two_weeks_ago_wtd_summary["ga4_FF_Lead_Event_Count"])

    # Ratio metrics for the filtered data last_year_same_iso_week_wtd_summary
    last_year_same_iso_week_wtd_cac = safe_divide(last_year_same_iso_week_wtd_summary['Cost'], last_year_same_iso_week_wtd_summary["Furnished Finder - GA4 (web) FF Purchase"])
    last_year_same_iso_week_wtd_cpl = safe_divide(last_year_same_iso_week_wtd_summary['Cost'], last_year_same_iso_week_wtd_summary["Furnished Finder - GA4 (web) FF Lead"])
    last_year_same_iso_week_wtd_cpc = safe_divide(last_year_same_iso_week_wtd_summary['Cost'], last_year_same_iso_week_wtd_summary['Clicks'])
    last_year_same_iso_week_wtd_lead_cvr = safe_divide(last_year_same_iso_week_wtd_summary["Furnished Finder - GA4 (web) FF Purchase"], last_year_same_iso_week_wtd_summary["Furnished Finder - GA4 (web) FF Lead"])
    last_year_same_iso_week_wtd_click_cvr = safe_divide(last_year_same_iso_week_wtd_summary["Furnished Finder - GA4 (web) FF Lead"], last_year_same_iso_week_wtd_summary['Clicks'])
    last_year_same_iso_week_wtd_cost_per_brsubmit = safe_divide(last_year_same_iso_week_wtd_summary['Cost'], last_year_same_iso_week_wtd_summary["Furnished Finder - GA4 (web) FF-BRSubmit"])
    last_year_same_iso_week_wtd_brsubmit_clicks_cvr = safe_divide(last_year_same_iso_week_wtd_summary["Furnished Finder - GA4 (web) FF-BRSubmit"], last_year_same_iso_week_wtd_summary['Clicks'])

    last_year_same_iso_week_wtd_ga4_cac = safe_divide(last_year_same_iso_week_wtd_summary['Cost'], last_year_same_iso_week_wtd_summary["ga4_FF_Purchase_Event_Count"])
    last_year_same_iso_week_wtd_ga4_lead_cvr = safe_divide(last_year_same_iso_week_wtd_summary["ga4_FF_Purchase_Event_Count"], last_year_same_iso_week_wtd_summary["ga4_FF_Lead_Event_Count"])

    # Ratio metrics for the filtered data avg_same_wtd_summary change to avg_same_wtd_26wk_summary
    avg_same_wtd_26wk_cac = safe_divide(avg_same_wtd_26wk_summary['Cost'], avg_same_wtd_26wk_summary["Furnished Finder - GA4 (web) FF Purchase"])
    avg_same_wtd_26wk_cpl = safe_divide(avg_same_wtd_26wk_summary['Cost'], avg_same_wtd_26wk_summary["Furnished Finder - GA4 (web) FF Lead"])
    avg_same_wtd_26wk_cpc = safe_divide(avg_same_wtd_26wk_summary['Cost'], avg_same_wtd_26wk_summary['Clicks'])
    avg_same_wtd_26wk_lead_cvr = safe_divide(avg_same_wtd_26wk_summary["Furnished Finder - GA4 (web) FF Purchase"], avg_same_wtd_26wk_summary["Furnished Finder - GA4 (web) FF Lead"])
    avg_same_wtd_26wk_click_cvr = safe_divide(avg_same_wtd_26wk_summary["Furnished Finder - GA4 (web) FF Lead"], avg_same_wtd_26wk_summary['Clicks'])
    avg_same_wtd_26wk_cost_per_brsubmit = safe_divide(avg_same_wtd_26wk_summary['Cost'], avg_same_wtd_26wk_summary["Furnished Finder - GA4 (web) FF-BRSubmit"])
    avg_same_wtd_26wk_brsubmit_clicks_cvr = safe_divide(avg_same_wtd_26wk_summary["Furnished Finder - GA4 (web) FF-BRSubmit"], avg_same_wtd_26wk_summary['Clicks'])

    avg_same_wtd_26wk_ga4_cac = safe_divide(avg_same_wtd_26wk_summary['Cost'], avg_same_wtd_26wk_summary["ga4_FF_Purchase_Event_Count"])
    avg_same_wtd_26wk_ga4_lead_cvr = safe_divide(avg_same_wtd_26wk_summary["ga4_FF_Purchase_Event_Count"], avg_same_wtd_26wk_summary["ga4_FF_Lead_Event_Count"])

    # Ratio metrics for the filtered data avg_same_weekday_summary change to avg_same_weekday_52wk_summary
    avg_same_wtd_52wk_cac = safe_divide(avg_same_wtd_52wk_summary['Cost'], avg_same_wtd_52wk_summary["Furnished Finder - GA4 (web) FF Purchase"])
    avg_same_wtd_52wk_cpl = safe_divide(avg_same_wtd_52wk_summary['Cost'], avg_same_wtd_52wk_summary["Furnished Finder - GA4 (web) FF Lead"])
    avg_same_wtd_52wk_cpc = safe_divide(avg_same_wtd_52wk_summary['Cost'], avg_same_wtd_52wk_summary['Clicks'])
    avg_same_wtd_52wk_lead_cvr = safe_divide(avg_same_wtd_52wk_summary["Furnished Finder - GA4 (web) FF Purchase"], avg_same_wtd_52wk_summary["Furnished Finder - GA4 (web) FF Lead"])
    avg_same_wtd_52wk_click_cvr = safe_divide(avg_same_wtd_52wk_summary["Furnished Finder - GA4 (web) FF Lead"], avg_same_wtd_52wk_summary['Clicks'])
    avg_same_wtd_52wk_cost_per_brsubmit = safe_divide(avg_same_wtd_52wk_summary['Cost'], avg_same_wtd_52wk_summary["Furnished Finder - GA4 (web) FF-BRSubmit"])
    avg_same_wtd_52wk_brsubmit_clicks_cvr = safe_divide(avg_same_wtd_52wk_summary["Furnished Finder - GA4 (web) FF-BRSubmit"], avg_same_wtd_52wk_summary['Clicks'])

    avg_same_wtd_52wk_ga4_cac = safe_divide(avg_same_wtd_52wk_summary['Cost'], avg_same_wtd_52wk_summary["ga4_FF_Purchase_Event_Count"])
    avg_same_wtd_52wk_ga4_lead_cvr = safe_divide(avg_same_wtd_52wk_summary["ga4_FF_Purchase_Event_Count"], avg_same_wtd_52wk_summary["ga4_FF_Lead_Event_Count"])

    # Ratio metrics for the filtered data mtd_summary
    mtd_cac = safe_divide(mtd_summary['Cost'], mtd_summary["Furnished Finder - GA4 (web) FF Purchase"])
    mtd_cpl = safe_divide(mtd_summary['Cost'], mtd_summary["Furnished Finder - GA4 (web) FF Lead"])
    mtd_cpc = safe_divide(mtd_summary['Cost'], mtd_summary['Clicks'])
    mtd_lead_cvr = safe_divide(mtd_summary["Furnished Finder - GA4 (web) FF Purchase"], mtd_summary["Furnished Finder - GA4 (web) FF Lead"])
    mtd_click_cvr = safe_divide(mtd_summary["Furnished Finder - GA4 (web) FF Lead"], mtd_summary['Clicks'])
    mtd_cost_per_brsubmit = safe_divide(mtd_summary['Cost'], mtd_summary["Furnished Finder - GA4 (web) FF-BRSubmit"])
    mtd_brsubmit_clicks_cvr = safe_divide(mtd_summary["Furnished Finder - GA4 (web) FF-BRSubmit"], mtd_summary['Clicks'])

    mtd_ga4_cac = safe_divide(mtd_summary['Cost'], mtd_summary["ga4_FF_Purchase_Event_Count"])
    mtd_ga4_lead_cvr = safe_divide(mtd_summary["ga4_FF_Purchase_Event_Count"], mtd_summary["ga4_FF_Lead_Event_Count"])

    # Ratio metrics for the filtered data last_month_mtd_summary
    last_month_mtd_cac = safe_divide(last_month_mtd_summary['Cost'], last_month_mtd_summary["Furnished Finder - GA4 (web) FF Purchase"])
    last_month_mtd_cpl = safe_divide(last_month_mtd_summary['Cost'], last_month_mtd_summary["Furnished Finder - GA4 (web) FF Lead"])
    last_month_mtd_cpc = safe_divide(last_month_mtd_summary['Cost'], last_month_mtd_summary['Clicks'])
    last_month_mtd_lead_cvr = safe_divide(last_month_mtd_summary["Furnished Finder - GA4 (web) FF Purchase"], last_month_mtd_summary["Furnished Finder - GA4 (web) FF Lead"])
    last_month_mtd_click_cvr = safe_divide(last_month_mtd_summary["Furnished Finder - GA4 (web) FF Lead"], last_month_mtd_summary['Clicks'])
    last_month_mtd_cost_per_brsubmit = safe_divide(last_month_mtd_summary['Cost'], last_month_mtd_summary["Furnished Finder - GA4 (web) FF-BRSubmit"])
    last_month_mtd_brsubmit_clicks_cvr = safe_divide(last_month_mtd_summary["Furnished Finder - GA4 (web) FF-BRSubmit"], last_month_mtd_summary['Clicks'])

    last_month_mtd_ga4_cac = safe_divide(last_month_mtd_summary['Cost'], last_month_mtd_summary["ga4_FF_Purchase_Event_Count"])
    last_month_mtd_ga4_lead_cvr = safe_divide(last_month_mtd_summary["ga4_FF_Purchase_Event_Count"], last_month_mtd_summary["ga4_FF_Lead_Event_Count"])

    # Ratio metrics for the filtered data two_months_ago_mtd_summary
    two_months_ago_mtd_cac = safe_divide(two_months_ago_mtd_summary['Cost'], two_months_ago_mtd_summary["Furnished Finder - GA4 (web) FF Purchase"])
    two_months_ago_mtd_cpl = safe_divide(two_months_ago_mtd_summary['Cost'], two_months_ago_mtd_summary["Furnished Finder - GA4 (web) FF Lead"])
    two_months_ago_mtd_cpc = safe_divide(two_months_ago_mtd_summary['Cost'], two_months_ago_mtd_summary['Clicks'])
    two_months_ago_mtd_lead_cvr = safe_divide(two_months_ago_mtd_summary["Furnished Finder - GA4 (web) FF Purchase"], two_months_ago_mtd_summary["Furnished Finder - GA4 (web) FF Lead"])
    two_months_ago_mtd_click_cvr = safe_divide(two_months_ago_mtd_summary["Furnished Finder - GA4 (web) FF Lead"], two_months_ago_mtd_summary['Clicks'])
    two_months_ago_mtd_cost_per_brsubmit = safe_divide(two_months_ago_mtd_summary['Cost'], two_months_ago_mtd_summary["Furnished Finder - GA4 (web) FF-BRSubmit"])
    two_months_ago_mtd_brsubmit_clicks_cvr = safe_divide(two_months_ago_mtd_summary["Furnished Finder - GA4 (web) FF-BRSubmit"], two_months_ago_mtd_summary['Clicks'])

    two_months_ago_mtd_ga4_cac = safe_divide(two_months_ago_mtd_summary['Cost'], two_months_ago_mtd_summary["ga4_FF_Purchase_Event_Count"])
    two_months_ago_mtd_ga4_lead_cvr = safe_divide(two_months_ago_mtd_summary["ga4_FF_Purchase_Event_Count"], two_months_ago_mtd_summary["ga4_FF_Lead_Event_Count"])

    # Ratio metrics for the filtered data last_year_month_mtd_summary
    last_year_mtd_cac = safe_divide(last_year_month_mtd_summary['Cost'], last_year_month_mtd_summary["Furnished Finder - GA4 (web) FF Purchase"])
    last_year_mtd_cpl = safe_divide(last_year_month_mtd_summary['Cost'], last_year_month_mtd_summary["Furnished Finder - GA4 (web) FF Lead"])
    last_year_mtd_cpc = safe_divide(last_year_month_mtd_summary['Cost'], last_year_month_mtd_summary['Clicks'])
    last_year_mtd_lead_cvr = safe_divide(last_year_month_mtd_summary["Furnished Finder - GA4 (web) FF Purchase"], last_year_month_mtd_summary["Furnished Finder - GA4 (web) FF Lead"])
    last_year_mtd_click_cvr = safe_divide(last_year_month_mtd_summary["Furnished Finder - GA4 (web) FF Lead"], last_year_month_mtd_summary['Clicks'])
    last_year_mtd_cost_per_brsubmit = safe_divide(last_year_month_mtd_summary['Cost'], last_year_month_mtd_summary["Furnished Finder - GA4 (web) FF-BRSubmit"])
    last_year_mtd_brsubmit_clicks_cvr = safe_divide(last_year_month_mtd_summary["Furnished Finder - GA4 (web) FF-BRSubmit"], last_year_month_mtd_summary['Clicks'])

    last_year_mtd_ga4_cac = safe_divide(last_year_month_mtd_summary['Cost'], last_year_month_mtd_summary["ga4_FF_Purchase_Event_Count"])
    last_year_mtd_ga4_lead_cvr = safe_divide(last_year_month_mtd_summary["ga4_FF_Purchase_Event_Count"], last_year_month_mtd_summary["ga4_FF_Lead_Event_Count"])


    # Calculate percentage changes using safe_divide
    percent_change_dd = safe_divide(yesterday_summary - day_before_summary, day_before_summary)
    percent_change_wow = safe_divide(yesterday_summary - last_week_summary, last_week_summary)
    percent_change_2wow = safe_divide(yesterday_summary - two_weeks_ago_summary, two_weeks_ago_summary)
    percent_change_avg_26wk = safe_divide(yesterday_summary - avg_same_weekday_26wk_summary, avg_same_weekday_26wk_summary)
    percent_change_avg_52wk = safe_divide(yesterday_summary - avg_same_weekday_52wk_summary, avg_same_weekday_52wk_summary)
    percent_change_last_year_iso_week = safe_divide(yesterday_summary - last_year_same_iso_week_summary, last_year_same_iso_week_summary)

    # Calculate WTD percentage changes using safe_divide
    percent_change_wtd = safe_divide(wtd_summary - last_week_wtd_summary, last_week_wtd_summary)
    percent_change_2w_wtd = safe_divide(wtd_summary - two_weeks_ago_wtd_summary, two_weeks_ago_wtd_summary)
    percent_change_ytd_wtd = safe_divide(wtd_summary - last_year_same_iso_week_wtd_summary, last_year_same_iso_week_wtd_summary)
    percent_change_avg_26wk_wtd = safe_divide(wtd_summary - avg_same_wtd_26wk_summary, avg_same_wtd_26wk_summary)
    percent_change_avg_52wk_wtd = safe_divide(wtd_summary - avg_same_wtd_52wk_summary, avg_same_wtd_52wk_summary)
    percent_change_last_year_iso_week = safe_divide(wtd_summary - last_year_same_iso_week_summary, last_year_same_iso_week_summary)

    # Calculate MTD percentage changes using safe_divide
    percent_change_last_month_mtd = safe_divide(mtd_summary - last_month_mtd_summary, last_month_mtd_summary)
    percent_change_2m_mtd = safe_divide(mtd_summary - two_months_ago_mtd_summary, two_months_ago_mtd_summary)
    percent_change_ytd_mtd = safe_divide(mtd_summary - last_year_month_mtd_summary, last_year_month_mtd_summary)
    
    # Calculate yesterday d/d percentage changes
    cac_dd = safe_divide((cac - day_before_cac), day_before_cac)
    cpl_dd = safe_divide((cpl - day_before_cpl), day_before_cpl)
    cpc_dd = safe_divide((cpc - day_before_cpc), day_before_cpc)
    lead_cvr_dd = safe_divide((lead_cvr - day_before_lead_cvr), day_before_lead_cvr)
    click_cvr_dd = safe_divide((click_cvr - day_before_click_cvr), day_before_click_cvr)

    # Calculate GA4 yesterday d/d percentage changes
    ga4_cac_dd = safe_divide((ga4_cac - day_before_ga4_cac), day_before_ga4_cac)
    ga4_lead_cvr_dd = safe_divide((ga4_lead_cvr - day_before_ga4_lead_cvr), day_before_ga4_lead_cvr)

    # Calculate yesterday WoW percentage changes
    cac_wow = safe_divide((cac - last_week_cac), last_week_cac)
    cpl_wow = safe_divide((cpl - last_week_cpl), last_week_cpl)
    cpc_wow = safe_divide((cpc - last_week_cpc), last_week_cpc)
    lead_cvr_wow = safe_divide((lead_cvr - last_week_lead_cvr), last_week_lead_cvr)
    click_cvr_wow = safe_divide((click_cvr - last_week_click_cvr), last_week_click_cvr)

    # Calculate GA4 yesterday WoW percentage changes
    ga4_cac_wow = safe_divide((ga4_cac - last_week_ga4_cac), last_week_ga4_cac)
    ga4_lead_cvr_wow = safe_divide((ga4_lead_cvr - last_week_ga4_lead_cvr), last_week_ga4_lead_cvr)

    # Calculate yesterday two_weeks_ago percentage changes
    cac_2wow = safe_divide((cac - two_weeks_ago_cac), two_weeks_ago_cac)
    cpl_2wow = safe_divide((cpl - two_weeks_ago_cpl), two_weeks_ago_cpl)
    cpc_2wow = safe_divide((cpc - two_weeks_ago_cpc), two_weeks_ago_cpc)
    lead_cvr_2wow = safe_divide((lead_cvr - two_weeks_ago_lead_cvr), two_weeks_ago_lead_cvr)
    click_cvr_2wow = safe_divide((click_cvr - two_weeks_ago_click_cvr), two_weeks_ago_click_cvr)

    # Calculate GA4 yesterday two_weeks_ago percentage changes
    ga4_cac_2wow = safe_divide((ga4_cac - two_weeks_ago_ga4_cac), two_weeks_ago_ga4_cac)
    ga4_lead_cvr_2wow = safe_divide((ga4_lead_cvr - two_weeks_ago_ga4_lead_cvr), two_weeks_ago_ga4_lead_cvr)

    # Calculate yesterday avg_same_weekday_26wk_wow percentage changes
    cac_avg_26wk_wow = safe_divide((cac - avg_same_weekday_26wk_cac), avg_same_weekday_26wk_cac)
    cpl_avg_26wk_wow = safe_divide((cpl - avg_same_weekday_26wk_cpl), avg_same_weekday_26wk_cpl)
    cpc_avg_26wk_wow = safe_divide((cpc - avg_same_weekday_26wk_cpc), avg_same_weekday_26wk_cpc)
    lead_cvr_avg_26wk_wow = safe_divide((lead_cvr - avg_same_weekday_26wk_lead_cvr), avg_same_weekday_26wk_lead_cvr)
    click_cvr_avg_26wk_wow = safe_divide((click_cvr - avg_same_weekday_26wk_click_cvr), avg_same_weekday_26wk_click_cvr)

    # Calculate GA4 yesterday avg_same_weekday_26wk_wow percentage changes
    ga4_cac_avg_26wk_wow = safe_divide((ga4_cac - avg_same_weekday_26wk_ga4_cac), avg_same_weekday_26wk_ga4_cac)
    ga4_lead_cvr_avg_26wk_wow = safe_divide((ga4_lead_cvr - avg_same_weekday_26wk_ga4_lead_cvr), avg_same_weekday_26wk_ga4_lead_cvr)

    # Calculate yesterday avg_same_weekday percentage changes
    cac_avg_52wk_wow = safe_divide((cac - avg_same_weekday_52wk_cac), avg_same_weekday_52wk_cac)
    cpl_avg_52wk_wow = safe_divide((cpl - avg_same_weekday_52wk_cpl), avg_same_weekday_52wk_cpl)
    cpc_avg_52wk_wow = safe_divide((cpc - avg_same_weekday_52wk_cpc), avg_same_weekday_52wk_cpc)
    lead_cvr_avg_52wk_wow = safe_divide((lead_cvr - avg_same_weekday_52wk_lead_cvr), avg_same_weekday_52wk_lead_cvr)
    click_cvr_avg_52wk_wow = safe_divide((click_cvr - avg_same_weekday_52wk_click_cvr), avg_same_weekday_52wk_click_cvr)

    # Calculate GA4 yesterday avg_same_weekday_52wk_wow percentage changes
    ga4_cac_avg_52wk_wow = safe_divide((ga4_cac - avg_same_weekday_52wk_ga4_cac), avg_same_weekday_52wk_ga4_cac)
    ga4_lead_cvr_avg_52wk_wow = safe_divide((ga4_lead_cvr - avg_same_weekday_52wk_ga4_lead_cvr), avg_same_weekday_52wk_ga4_lead_cvr)

    # Calculate yesterday last_year_same_iso_week percentage changes
    cac_lywow = safe_divide((cac - last_year_same_iso_week_cac), last_year_same_iso_week_cac)
    cpl_lywow = safe_divide((cpl - last_year_same_iso_week_cpl), last_year_same_iso_week_cpl)
    cpc_lywow = safe_divide((cpc - last_year_same_iso_week_cpc), last_year_same_iso_week_cpc)
    lead_cvr_lywow = safe_divide((lead_cvr - last_year_same_iso_week_lead_cvr), last_year_same_iso_week_lead_cvr)
    click_cvr_lywow = safe_divide((click_cvr - last_year_same_iso_week_click_cvr), last_year_same_iso_week_click_cvr)

    # Calculate GA4 yesterday last_year_same_iso_week percentage changes
    ga4_cac_lywow = safe_divide((ga4_cac - last_year_same_iso_week_ga4_cac), last_year_same_iso_week_ga4_cac)
    ga4_lead_cvr_lywow = safe_divide((ga4_lead_cvr - last_year_same_iso_week_ga4_lead_cvr), last_year_same_iso_week_ga4_lead_cvr)

    # Calculate WTD WoW percentage changes
    cac_wtdwow = safe_divide((wtd_cac - last_week_wtd_cac), last_week_wtd_cac)
    cpl_wtdwow = safe_divide((wtd_cpl - last_week_wtd_cpl), last_week_wtd_cpl)
    cpc_wtdwow = safe_divide((wtd_cpc - last_week_wtd_cpc), last_week_wtd_cpc)
    lead_cvr_wtdwow = safe_divide((wtd_lead_cvr - last_week_wtd_lead_cvr), last_week_wtd_lead_cvr)
    click_cvr_wtdwow = safe_divide((wtd_click_cvr - last_week_wtd_click_cvr), last_week_wtd_click_cvr)

    # Calculate GA4 WTD WoW percentage changes
    ga4_cac_wtdwow = safe_divide((wtd_ga4_cac - last_week_wtd_ga4_cac), last_week_wtd_ga4_cac)
    ga4_lead_cvr_wtdwow = safe_divide((wtd_ga4_lead_cvr - last_week_wtd_ga4_lead_cvr), last_week_wtd_ga4_lead_cvr)

    # Calculate WTD Wo2W percentage changes
    cac_wtdwo2w = safe_divide((wtd_cac - two_weeks_ago_wtd_cac), two_weeks_ago_wtd_cac)
    cpl_wtdwo2w = safe_divide((wtd_cpl - two_weeks_ago_wtd_cpl), two_weeks_ago_wtd_cpl)
    cpc_wtdwo2w = safe_divide((wtd_cpc - two_weeks_ago_wtd_cpc), two_weeks_ago_wtd_cpc)
    lead_cvr_wtdwo2w = safe_divide((wtd_lead_cvr - two_weeks_ago_wtd_lead_cvr), two_weeks_ago_wtd_lead_cvr)
    click_cvr_wtdwo2w = safe_divide((wtd_click_cvr - two_weeks_ago_wtd_click_cvr), two_weeks_ago_wtd_click_cvr)

    # Calculate GA4 WTD Wo2W percentage changes
    ga4_cac_wtdwo2w = safe_divide((wtd_ga4_cac - two_weeks_ago_wtd_ga4_cac), two_weeks_ago_wtd_ga4_cac)
    ga4_lead_cvr_wtdwo2w = safe_divide((wtd_ga4_lead_cvr - two_weeks_ago_wtd_ga4_lead_cvr), two_weeks_ago_wtd_ga4_lead_cvr)

    # Calculate WTD percentage changes last_year_same_iso_week_wtd
    cac_lywtdwow = safe_divide((wtd_cac - last_year_same_iso_week_wtd_cac), last_year_same_iso_week_wtd_cac)
    cpl_lywtdwow = safe_divide((wtd_cpl - last_year_same_iso_week_wtd_cpl), last_year_same_iso_week_wtd_cpl)
    cpc_lywtdwow = safe_divide((wtd_cpc - last_year_same_iso_week_wtd_cpc), last_year_same_iso_week_wtd_cpc)
    lead_cvr_lywtdwow = safe_divide((wtd_lead_cvr - last_year_same_iso_week_wtd_lead_cvr), last_year_same_iso_week_wtd_lead_cvr)
    click_cvr_lywtdwow = safe_divide((wtd_click_cvr - last_year_same_iso_week_wtd_click_cvr), last_year_same_iso_week_wtd_click_cvr)

    # Calculate GA4 WTD percentage changes last_year_same_iso_week_wtd
    ga4_cac_lywtdwow = safe_divide((wtd_ga4_cac - last_year_same_iso_week_wtd_ga4_cac), last_year_same_iso_week_wtd_ga4_cac)
    ga4_lead_cvr_lywtdwow = safe_divide((wtd_ga4_lead_cvr - last_year_same_iso_week_wtd_ga4_lead_cvr), last_year_same_iso_week_wtd_ga4_lead_cvr)

    # Calculate yesterday avg_same_weekday_26wk_wow_wtd percentage changes
    cac_avg_26wk_wow_wtd = safe_divide((cac - avg_same_wtd_26wk_cac), avg_same_wtd_26wk_cac)
    cpl_avg_26wk_wow_wtd = safe_divide((cpl - avg_same_wtd_26wk_cpl), avg_same_wtd_26wk_cpl)
    cpc_avg_26wk_wow_wtd = safe_divide((cpc - avg_same_wtd_26wk_cpc), avg_same_wtd_26wk_cpc)
    lead_cvr_avg_26wk_wow_wtd = safe_divide((lead_cvr - avg_same_wtd_26wk_lead_cvr), avg_same_wtd_26wk_lead_cvr)
    click_cvr_avg_26wk_wow_wtd = safe_divide((click_cvr - avg_same_wtd_26wk_click_cvr), avg_same_wtd_26wk_click_cvr)

    # Calculate GA4 yesterday avg_same_weekday_26wk_wow_wtd percentage changes
    ga4_cac_avg_26wk_wow_wtd = safe_divide((ga4_cac - avg_same_wtd_26wk_ga4_cac), avg_same_wtd_26wk_ga4_cac)
    ga4_lead_cvr_avg_26wk_wow_wtd = safe_divide((ga4_lead_cvr - avg_same_wtd_26wk_ga4_lead_cvr), avg_same_wtd_26wk_ga4_lead_cvr)

    # Calculate yesterday avg_same_weekday_wtd percentage changes
    cac_avg_52wk_wow_wtd = safe_divide((cac - avg_same_wtd_52wk_cac), avg_same_wtd_52wk_cac)
    cpl_avg_52wk_wow_wtd = safe_divide((cpl - avg_same_wtd_52wk_cpl), avg_same_wtd_52wk_cpl)
    cpc_avg_52wk_wow_wtd = safe_divide((cpc - avg_same_wtd_52wk_cpc), avg_same_wtd_52wk_cpc)
    lead_cvr_avg_52wk_wow_wtd = safe_divide((lead_cvr - avg_same_wtd_52wk_lead_cvr), avg_same_wtd_52wk_lead_cvr)
    click_cvr_avg_52wk_wow_wtd = safe_divide((click_cvr - avg_same_wtd_52wk_click_cvr), avg_same_wtd_52wk_click_cvr)

    # Calculate GA4 yesterday avg_same_weekday_wtd percentage changes
    ga4_cac_avg_52wk_wow_wtd = safe_divide((ga4_cac - avg_same_wtd_52wk_ga4_cac), avg_same_wtd_52wk_ga4_cac)
    ga4_lead_cvr_avg_52wk_wow_wtd = safe_divide((ga4_lead_cvr - avg_same_wtd_52wk_ga4_lead_cvr), avg_same_wtd_52wk_ga4_lead_cvr)

    # Calculate MTD percentage changes last_month_mtd_summary
    cac_mtdmom = safe_divide((mtd_cac - last_month_mtd_cac), last_month_mtd_cac)
    cpl_mtdmom = safe_divide((mtd_cpl - last_month_mtd_cpl), last_month_mtd_cpl)
    cpc_mtdmom = safe_divide((mtd_cpc - last_month_mtd_cpc), last_month_mtd_cpc)
    lead_cvr_mtdmom = safe_divide((mtd_lead_cvr - last_month_mtd_lead_cvr), last_month_mtd_lead_cvr)
    click_cvr_mtdmom = safe_divide((mtd_click_cvr - last_month_mtd_click_cvr), last_month_mtd_click_cvr)

    # Calculate GA4 MTD percentage changes last_month_mtd_summary
    ga4_cac_mtdmom = safe_divide((mtd_ga4_cac - last_month_mtd_ga4_cac), last_month_mtd_ga4_cac)
    ga4_lead_cvr_mtdmom = safe_divide((mtd_ga4_lead_cvr - last_month_mtd_ga4_lead_cvr), last_month_mtd_ga4_lead_cvr)

    # Calculate MTD percentage changes two_months_ago_mtd
    cac_mtdmo2m = safe_divide((mtd_cac - two_months_ago_mtd_cac), two_months_ago_mtd_cac)
    cpl_mtdmo2m = safe_divide((mtd_cpl - two_months_ago_mtd_cpl), two_months_ago_mtd_cpl)
    cpc_mtdmo2m = safe_divide((mtd_cpc - two_months_ago_mtd_cpc), two_months_ago_mtd_cpc)
    lead_cvr_mtdmo2m = safe_divide((mtd_lead_cvr - two_months_ago_mtd_lead_cvr), two_months_ago_mtd_lead_cvr)
    click_cvr_mtdmo2m = safe_divide((mtd_click_cvr - two_months_ago_mtd_click_cvr), two_months_ago_mtd_click_cvr)

    # Calculate GA4 MTD percentage changes two_months_ago_mtd
    ga4_cac_mtdmo2m = safe_divide((mtd_ga4_cac - two_months_ago_mtd_ga4_cac), two_months_ago_mtd_ga4_cac)
    ga4_lead_cvr_mtdmo2m = safe_divide((mtd_ga4_lead_cvr - two_months_ago_mtd_ga4_lead_cvr), two_months_ago_mtd_ga4_lead_cvr)

    # Calculate MTD percentage changes last_year_month_mtd
    cac_mtdmolym = safe_divide((mtd_cac - last_year_mtd_cac), last_year_mtd_cac)
    cpl_mtdmolym = safe_divide((mtd_cpl - last_year_mtd_cpl), last_year_mtd_cpl)
    cpc_mtdmolym = safe_divide((mtd_cpc - last_year_mtd_cpc), last_year_mtd_cpc)
    lead_cvr_mtdmolym = safe_divide((mtd_lead_cvr - last_year_mtd_lead_cvr), last_year_mtd_lead_cvr)
    click_cvr_mtdmolym = safe_divide((mtd_click_cvr - last_year_mtd_click_cvr), last_year_mtd_click_cvr)

    # Calculate GA4 MTD percentage changes last_year_month_mtd
    ga4_cac_mtdmolym = safe_divide((mtd_ga4_cac - last_year_mtd_ga4_cac), last_year_mtd_ga4_cac)
    ga4_lead_cvr_mtdmolym = safe_divide((mtd_ga4_lead_cvr - last_year_mtd_ga4_lead_cvr), last_year_mtd_ga4_lead_cvr)


    # Create summary DataFrame
    summary_df = pd.DataFrame({
        "Metric": metrics,
        "Yesterday": yesterday_summary.values,
        "Day Before": day_before_summary.values,
        "d/d %": percent_change_dd,
        "d w/w %": percent_change_wow,
        "d w/2w %": percent_change_2wow,
        "d Avg 26w %": percent_change_avg_26wk,
        "d Avg 52w %": percent_change_avg_52wk,
        "y/y (iso week) %": percent_change_last_year_iso_week,
        "WTD": wtd_summary.values,
        "Last Week WTD": last_week_wtd_summary.values,
        "WTD w/w %": percent_change_wtd,
        "WTD w/2w %": percent_change_2w_wtd,
        "WTD Avg 26w %": percent_change_avg_26wk_wtd,
        "WTD Avg 52w %": percent_change_avg_52wk_wtd,
        "WTD y/y (iso week) %": percent_change_ytd_wtd,
        "MTD": mtd_summary.values,
        "Last Month MTD": last_month_mtd_summary.values,
        "MTD m/m %": percent_change_last_month_mtd,
        "MTD 2m/2m %": percent_change_2m_mtd,
        "MTD y/y (iso month) %": percent_change_ytd_mtd
    })

    
    # Define a helper function to ensure safe values
    def safe_value(value):
        return value if pd.notna(value) else 0

    # Create the summary_ratio_df with safe values
    summary_ratio_df = pd.DataFrame({
        "Metric": [
            'CAC', 'CPL', 'CPC', 'Purchase : Lead CVR', 'Lead : Click CVR', 'Cost Per BRSubmit', 'BRSubmit : Click CVR',
            'GA4 CAC', 'GA4 Purchase : Lead CVR'
        ],
        "Yesterday": [
            safe_value(cac), safe_value(cpl), safe_value(cpc), safe_value(lead_cvr), safe_value(click_cvr),
            safe_value(cost_per_brsubmit), safe_value(brsubmit_clicks_cvr),
            safe_value(ga4_cac), safe_value(ga4_lead_cvr)
        ],
        "Day Before": [
            safe_value(day_before_cac), safe_value(day_before_cpl), safe_value(day_before_cpc), safe_value(day_before_lead_cvr),
            safe_value(day_before_click_cvr), safe_value(day_before_cost_per_brsubmit), safe_value(day_before_brsubmit_clicks_cvr),
            safe_value(day_before_ga4_cac), safe_value(day_before_ga4_lead_cvr)
        ],
        "d/d %": [
            safe_value(cac_dd), safe_value(cpl_dd), safe_value(cpc_dd), safe_value(lead_cvr_dd), safe_value(click_cvr_dd),
            safe_divide(cost_per_brsubmit - day_before_cost_per_brsubmit, day_before_cost_per_brsubmit),
            safe_divide(brsubmit_clicks_cvr - day_before_brsubmit_clicks_cvr, day_before_brsubmit_clicks_cvr),
            safe_value(ga4_cac_dd), safe_value(ga4_lead_cvr_dd)
        ],
        "d w/w %": [
            safe_value(cac_wow), safe_value(cpl_wow), safe_value(cpc_wow), safe_value(lead_cvr_wow), safe_value(click_cvr_wow),
            safe_divide(cost_per_brsubmit - last_week_cost_per_brsubmit, last_week_cost_per_brsubmit),
            safe_divide(brsubmit_clicks_cvr - last_week_brsubmit_clicks_cvr, last_week_brsubmit_clicks_cvr),
            safe_value(ga4_cac_wow), safe_value(ga4_lead_cvr_wow)
        ],
        "d w/2w %": [
            safe_value(cac_2wow), safe_value(cpl_2wow), safe_value(cpc_2wow), safe_value(lead_cvr_2wow), safe_value(click_cvr_2wow),
            safe_divide(cost_per_brsubmit - two_weeks_ago_cost_per_brsubmit, two_weeks_ago_cost_per_brsubmit),
            safe_divide(brsubmit_clicks_cvr - two_weeks_ago_brsubmit_clicks_cvr, two_weeks_ago_brsubmit_clicks_cvr),
            safe_value(ga4_cac_2wow), safe_value(ga4_lead_cvr_2wow)
        ],
        "d Avg 26w %": [
            safe_value(cac_avg_26wk_wow), safe_value(cpl_avg_26wk_wow), safe_value(cpc_avg_26wk_wow), safe_value(lead_cvr_avg_26wk_wow),
            safe_value(click_cvr_avg_26wk_wow), safe_divide(cost_per_brsubmit - avg_same_weekday_26wk_cost_per_brsubmit,
                                                            avg_same_weekday_26wk_cost_per_brsubmit),
            safe_divide(brsubmit_clicks_cvr - avg_same_weekday_26wk_brsubmit_clicks_cvr, avg_same_weekday_26wk_brsubmit_clicks_cvr),
            safe_value(ga4_cac_avg_26wk_wow), safe_value(ga4_lead_cvr_avg_26wk_wow)
        ],
        "d Avg 52w %": [
            safe_value(cac_avg_52wk_wow), safe_value(cpl_avg_52wk_wow), safe_value(cpc_avg_52wk_wow), safe_value(lead_cvr_avg_52wk_wow),
            safe_value(click_cvr_avg_52wk_wow), safe_divide(cost_per_brsubmit - avg_same_weekday_52wk_cost_per_brsubmit,
                                                            avg_same_weekday_52wk_cost_per_brsubmit),
            safe_divide(brsubmit_clicks_cvr - avg_same_weekday_52wk_brsubmit_clicks_cvr, avg_same_weekday_52wk_brsubmit_clicks_cvr),
            safe_value(ga4_cac_avg_52wk_wow), safe_value(ga4_lead_cvr_avg_52wk_wow)
        ],
        "y/y (iso week) %": [
            safe_value(cac_lywow), safe_value(cpl_lywow), safe_value(cpc_lywow), safe_value(lead_cvr_lywow), safe_value(click_cvr_lywow),
            safe_divide(cost_per_brsubmit - last_year_same_iso_week_cost_per_brsubmit, last_year_same_iso_week_cost_per_brsubmit),
            safe_divide(brsubmit_clicks_cvr - last_year_same_iso_week_brsubmit_clicks_cvr, last_year_same_iso_week_brsubmit_clicks_cvr),
            safe_value(ga4_cac_lywow), safe_value(ga4_lead_cvr_lywow)
        ],
        "WTD": [
            safe_value(wtd_cac), safe_value(wtd_cpl), safe_value(wtd_cpc), safe_value(wtd_lead_cvr), safe_value(wtd_click_cvr),
            safe_value(wtd_cost_per_brsubmit), safe_value(wtd_brsubmit_clicks_cvr),
            safe_value(wtd_ga4_cac), safe_value(wtd_ga4_lead_cvr)
        ],
        "Last Week WTD": [
            safe_value(last_week_wtd_cac), safe_value(last_week_wtd_cpl), safe_value(last_week_wtd_cpc), safe_value(last_week_wtd_lead_cvr),
            safe_value(last_week_wtd_click_cvr), safe_value(last_week_wtd_cost_per_brsubmit), safe_value(last_week_wtd_brsubmit_clicks_cvr),
            safe_value(last_week_wtd_ga4_cac), safe_value(last_week_wtd_ga4_lead_cvr)
        ],
        "WTD w/w %": [
            safe_value(cac_wtdwow), safe_value(cpl_wtdwow), safe_value(cpc_wtdwow), safe_value(lead_cvr_wtdwow), safe_value(click_cvr_wtdwow),
            safe_divide(wtd_cost_per_brsubmit - last_week_wtd_cost_per_brsubmit, last_week_wtd_cost_per_brsubmit),
            safe_divide(wtd_brsubmit_clicks_cvr - last_week_wtd_brsubmit_clicks_cvr, last_week_wtd_brsubmit_clicks_cvr),
            safe_value(ga4_cac_wtdwow), safe_value(ga4_lead_cvr_wtdwow)
        ],
        "WTD w/2w %": [
            safe_value(cac_wtdwo2w), safe_value(cpl_wtdwo2w), safe_value(cpc_wtdwo2w), safe_value(lead_cvr_wtdwo2w), safe_value(click_cvr_wtdwo2w),
            safe_divide(wtd_cost_per_brsubmit - two_weeks_ago_wtd_cost_per_brsubmit, two_weeks_ago_wtd_cost_per_brsubmit),
            safe_divide(wtd_brsubmit_clicks_cvr - two_weeks_ago_wtd_brsubmit_clicks_cvr, two_weeks_ago_wtd_brsubmit_clicks_cvr),
            safe_value(ga4_cac_wtdwo2w), safe_value(ga4_lead_cvr_wtdwo2w)
        ],
        "WTD Avg 26w %": [
            safe_value(cac_avg_26wk_wow_wtd), safe_value(cpl_avg_26wk_wow_wtd), safe_value(cpc_avg_26wk_wow_wtd), safe_value(lead_cvr_avg_26wk_wow_wtd),
            safe_value(click_cvr_avg_26wk_wow_wtd), safe_divide(wtd_cost_per_brsubmit - avg_same_wtd_26wk_cost_per_brsubmit,
                                                                avg_same_wtd_26wk_cost_per_brsubmit),
            safe_divide(wtd_brsubmit_clicks_cvr - avg_same_wtd_26wk_brsubmit_clicks_cvr, avg_same_wtd_26wk_brsubmit_clicks_cvr),
            safe_value(ga4_cac_avg_26wk_wow_wtd), safe_value(ga4_lead_cvr_avg_26wk_wow_wtd)
        ],
        "WTD Avg 52w %": [
            safe_value(cac_avg_52wk_wow_wtd), safe_value(cpl_avg_52wk_wow_wtd), safe_value(cpc_avg_52wk_wow_wtd), safe_value(lead_cvr_avg_52wk_wow_wtd),
            safe_value(click_cvr_avg_52wk_wow_wtd), safe_divide(wtd_cost_per_brsubmit - avg_same_wtd_52wk_cost_per_brsubmit,
                                                                avg_same_wtd_52wk_cost_per_brsubmit),
            safe_divide(wtd_brsubmit_clicks_cvr - avg_same_wtd_52wk_brsubmit_clicks_cvr, avg_same_wtd_52wk_brsubmit_clicks_cvr),
            safe_value(ga4_cac_avg_52wk_wow_wtd), safe_value(ga4_lead_cvr_avg_52wk_wow_wtd)
        ],
        "WTD y/y (iso week) %": [
            safe_value(cac_lywtdwow), safe_value(cpl_lywtdwow), safe_value(cpc_lywtdwow), safe_value(lead_cvr_lywtdwow), safe_value(click_cvr_lywtdwow),
            safe_divide(wtd_cost_per_brsubmit - last_year_same_iso_week_wtd_cost_per_brsubmit, last_year_same_iso_week_wtd_cost_per_brsubmit),
            safe_divide(wtd_brsubmit_clicks_cvr - last_year_same_iso_week_wtd_brsubmit_clicks_cvr, last_year_same_iso_week_wtd_brsubmit_clicks_cvr),
            safe_value(ga4_cac_lywtdwow), safe_value(ga4_lead_cvr_lywtdwow)
        ],
        "MTD": [
            safe_value(mtd_cac), safe_value(mtd_cpl), safe_value(mtd_cpc), safe_value(mtd_lead_cvr), safe_value(mtd_click_cvr),
            safe_value(mtd_cost_per_brsubmit), safe_value(mtd_brsubmit_clicks_cvr),
            safe_value(mtd_ga4_cac), safe_value(mtd_ga4_lead_cvr)
        ],
        "Last Month MTD": [
            safe_value(last_month_mtd_cac), safe_value(last_month_mtd_cpl), safe_value(last_month_mtd_cpc), safe_value(last_month_mtd_lead_cvr),
            safe_value(last_month_mtd_click_cvr), safe_value(last_month_mtd_cost_per_brsubmit), safe_value(last_month_mtd_brsubmit_clicks_cvr),
            safe_value(last_month_mtd_ga4_cac), safe_value(last_month_mtd_ga4_lead_cvr)
        ],
        "MTD m/m %": [
            safe_value(cac_mtdmom), safe_value(cpl_mtdmom), safe_value(cpc_mtdmom), safe_value(lead_cvr_mtdmom), safe_value(click_cvr_mtdmom),
            safe_divide(mtd_cost_per_brsubmit - last_month_mtd_cost_per_brsubmit, last_month_mtd_cost_per_brsubmit),
            safe_divide(mtd_brsubmit_clicks_cvr - last_month_mtd_brsubmit_clicks_cvr, last_month_mtd_brsubmit_clicks_cvr),
            safe_value(ga4_cac_mtdmom), safe_value(ga4_lead_cvr_mtdmom)
        ],
        "MTD 2m/2m %": [
            safe_value(cac_mtdmo2m), safe_value(cpl_mtdmo2m), safe_value(cpc_mtdmo2m), safe_value(lead_cvr_mtdmo2m), safe_value(click_cvr_mtdmo2m),
            safe_divide(mtd_cost_per_brsubmit - two_months_ago_mtd_cost_per_brsubmit, two_months_ago_mtd_cost_per_brsubmit),
            safe_divide(mtd_brsubmit_clicks_cvr - two_months_ago_mtd_brsubmit_clicks_cvr, two_months_ago_mtd_brsubmit_clicks_cvr),
            safe_value(ga4_cac_mtdmo2m), safe_value(ga4_lead_cvr_mtdmo2m)
        ],
        "MTD y/y (iso month) %": [
            safe_value(cac_mtdmolym), safe_value(cpl_mtdmolym), safe_value(cpc_mtdmolym), safe_value(lead_cvr_mtdmolym), safe_value(click_cvr_mtdmolym),
            safe_divide(mtd_cost_per_brsubmit - last_year_mtd_cost_per_brsubmit, last_year_mtd_cost_per_brsubmit),
            safe_divide(mtd_brsubmit_clicks_cvr - last_year_mtd_brsubmit_clicks_cvr, last_year_mtd_brsubmit_clicks_cvr),
            safe_value(ga4_cac_mtdmolym), safe_value(ga4_lead_cvr_mtdmolym)
        ]
    })

    union_df = pd.concat([summary_df, summary_ratio_df])

    print(union_df)
    return union_df


# Replace with your actual customer ID
CUSTOMER_ID = '7554573980' # TNH: '7731032510' ; FF: '7554573980'

# Define tenant campaign names
tenant_campaigns = ["FF - Search - Brand - Broad", "Search - Tenants"]

# Get campaign data
campaign_data = get_campaign_data(client, CUSTOMER_ID)

# Get conversion action data
conversion_action_data = get_conversion_action_data(client, CUSTOMER_ID)

# -----------------------------------------------------------------------#
# --------------------Load & Clean GA4 Events Data-----------------------#
# -----------------------------------------------------------------------#

# Load the GA4 events data
ga4_events_data = pd.read_csv('ga4-reports/output/ga4_events_data.csv')

# Filter data where 'Source' is 'google' and 'Medium' is 'cpc'
ga4_events_data = ga4_events_data[(ga4_events_data['Source'] == 'google') & (ga4_events_data['Medium'] == 'cpc')]

# Exclude rows where 'Campaign' contains 'Tenants' or 'Display'
# ga4_events_data = ga4_events_data[~ga4_events_data['Campaign'].str.contains('Tenants', case=False)]
ga4_events_data = ga4_events_data[~ga4_events_data['Campaign'].str.contains('Display', case=False)]

# Convert 'Date' column from string to datetime
ga4_events_data['Date'] = pd.to_datetime(ga4_events_data['Date'], format='%Y%m%d')

# Print to verify the changes
print(ga4_events_data.head())

# Aggregate the relevant columns and rename them
aggregated_ga4_data = ga4_events_data.groupby(['Date', 'Campaign', 'Device']).agg({
    'FF_Purchase_Event_Count': 'sum',
    'FF_Lead_Event_Count': 'sum',
    'FF_BRSubmit_Event_Count': 'sum',
    'FF_DMSubmit_Event_Count': 'sum',
    'FF_PhoneGet_Event_Count': 'sum'
}).reset_index()

# Rename columns to include 'ga4_' prefix
aggregated_ga4_data.rename(columns={
    'FF_Purchase_Event_Count': 'ga4_FF_Purchase_Event_Count',
    'FF_Lead_Event_Count': 'ga4_FF_Lead_Event_Count',
    'FF_BRSubmit_Event_Count': 'ga4_FF_BRSubmit_Event_Count',
    'FF_DMSubmit_Event_Count': 'ga4_FF_DMSubmit_Event_Count',
    'FF_PhoneGet_Event_Count': 'ga4_FF_PhoneGet_Event_Count'
}, inplace=True)

# Add a 'Network' column with NaN values to the aggregated_ga4_data
aggregated_ga4_data['Network'] = pd.NA

# Print to verify the changes
print(aggregated_ga4_data.head())


# ---------------------------------------------------------------------- #
# ---------------------Merging Google Ads & GA4 Data-------------------- #
# ---------------------------------------------------------------------- #

# Get clicks, impressions, and spend data
clicks_impressions_spend_data = get_clicks_impressions_spend_data(client, CUSTOMER_ID)

if campaign_data is not None and conversion_action_data is not None:
    # Extract numerical Conversion Action ID from the string format in campaign data
    campaign_data["Conversion Action ID"] = campaign_data["Conversion Action ID"].apply(extract_conversion_action_id)
    
    # Ensure both columns are of the same type (string in this case)
    campaign_data["Conversion Action ID"] = campaign_data["Conversion Action ID"].astype(str)
    conversion_action_data["Conversion Action ID"] = conversion_action_data["Conversion Action ID"].astype(str)
    
    # Merge the two dataframes on Conversion Action ID
    merged_data = pd.merge(campaign_data, conversion_action_data, on="Conversion Action ID", how="left")

    # Pivot the data to have individual columns for each conversion action name
    pivot_data = merged_data.pivot_table(index=["Date", "Campaign ID", "Device", "Network"],
                                         columns="Conversion Action Name",
                                         values="Conversions",
                                         aggfunc="sum",
                                         fill_value=0).reset_index()
    
    # Flatten the column headers
    pivot_data.columns.name = None

else:
    print("Failed to retrieve data.")

if clicks_impressions_spend_data is not None:
    # Merge the pivoted conversion data with the clicks, impressions, and spend data
    final_data = pd.merge(pivot_data, clicks_impressions_spend_data, on=["Date", "Campaign ID", "Device", "Network"], how="right")

    # Filter out campaigns containing "tenant" or "display"
    # final_data = final_data[~final_data['Campaign Name'].str.contains('Tenants', case=False)]
    final_data = final_data[~final_data['Campaign Name'].str.contains('Display', case=False)]
    
    # Include only "SEARCH" or "SEARCH_PARTNERS" networks
    final_data = final_data[final_data['Network'].str.contains('SEARCH|SEARCH_PARTNERS', case=False)]

    # Save final data to Excel
    save_to_excel(final_data, 'reports/output/final_data.xlsx') #tnhfinal_data.xlsx'
    print("Data successfully saved to 'final_data.xlsx'")
    
    # ------------------------ Union between final_data and aggregated_ga4_data ---------------------- #

    # Convert 'Date' in both dataframes to the same type if needed
    aggregated_ga4_data['Date'] = pd.to_datetime(aggregated_ga4_data['Date'], format='%Y%m%d')
    final_data['Date'] = pd.to_datetime(final_data['Date'])

    # Rename 'Campaign' column in aggregated_ga4_data to match 'Campaign Name' in final_data
    aggregated_ga4_data.rename(columns={'Campaign': 'Campaign Name'}, inplace=True)

    # Ensure 'Device' columns match in case of case sensitivity issues
    final_data['Device'] = final_data['Device'].str.upper()
    aggregated_ga4_data['Device'] = aggregated_ga4_data['Device'].str.upper()

    # Add 'Network' column with NaN values to aggregated_ga4_data
    aggregated_ga4_data['Network'] = pd.NA

    # Ensure all columns exist in both dataframes before concatenating
    for col in ['ga4_FF_Purchase_Event_Count', 'ga4_FF_Lead_Event_Count', 'ga4_FF_BRSubmit_Event_Count', 'ga4_FF_DMSubmit_Event_Count', 'ga4_FF_PhoneGet_Event_Count']:
        if col not in final_data.columns:
            final_data[col] = 0

    # Concatenate the dataframes vertically
    final_combined_data = pd.concat([final_data, aggregated_ga4_data], ignore_index=True)

    # Add the new calculated columns based on the dollar values
    final_combined_data['ga4_FF_Purchase_Event_Value'] = final_combined_data['ga4_FF_Purchase_Event_Count'] * 171
    final_combined_data['ga4_FF_BRSubmit_Event_Value'] = final_combined_data['ga4_FF_BRSubmit_Event_Count'] * 5
    final_combined_data['ga4_FF_DMSubmit_Event_Value'] = final_combined_data['ga4_FF_DMSubmit_Event_Count'] * 1.75
    final_combined_data['ga4_FF_PhoneGet_Event_Value'] = final_combined_data['ga4_FF_PhoneGet_Event_Count'] * 0.75

    # Save final combined data to CSV to inspect
    final_combined_data.to_csv('reports/output/final_combined_data.csv', index=False)

    # Print to verify the merge and new columns
    print(final_combined_data.head())

    
    # Loop through the last 90 days
    date_range = [datetime.today() - timedelta(days=x) for x in range(1, 90)]
    all_summaries = {}
    
    # Calculate and print the summary metrics for the last 90 days
    summary_data = {}
    for i in range(90):
        date = datetime.today() - timedelta(days=i)
        formatted_date = date.strftime('%Y%m%d')
        
        # Calculate summaries for overall data
        overall_summary = calculate_summary_metrics(final_combined_data, date)

        # Calculate summaries for brand and non-brand campaigns
        brand_campaigns = final_combined_data[final_combined_data['Campaign Name'].str.contains("Brand", case=False)]
        brand_summary = calculate_summary_metrics(brand_campaigns, date)
        non_brand_campaigns = final_combined_data[~final_combined_data['Campaign Name'].str.contains("Brand", case=False)]
        non_brand_summary = calculate_summary_metrics(non_brand_campaigns, date)
        
        # Calculate summaries by device for Brand Summary
        brand_desktop = brand_campaigns[brand_campaigns['Device'] == 'DESKTOP']
        brand_mobile = brand_campaigns[brand_campaigns['Device'] == 'MOBILE']
        brand_desktop_summary = calculate_summary_metrics(brand_desktop, date)
        brand_mobile_summary = calculate_summary_metrics(brand_mobile, date)
        
        # Calculate summaries by device for Non-Brand Summary
        non_brand_desktop = non_brand_campaigns[non_brand_campaigns['Device'] == 'DESKTOP']
        non_brand_mobile = non_brand_campaigns[non_brand_campaigns['Device'] == 'MOBILE']
        non_brand_desktop_summary = calculate_summary_metrics(non_brand_desktop, date)
        non_brand_mobile_summary = calculate_summary_metrics(non_brand_mobile, date)

        # Combine all summaries
        all_summaries[date.strftime('%Y%m%d')] = {
            "Google Overall Summary": overall_summary,
            "Google Brand Summary": brand_summary,
            "Google Non-Brand Summary": non_brand_summary,
            "Google Brand Desktop Summary": brand_desktop_summary,
            "Google Brand Mobile Summary": brand_mobile_summary,
            "Google Non-Brand Desktop Summary": non_brand_desktop_summary,
            "Google Non-Brand Mobile Summary": non_brand_mobile_summary,
        }

    # Save all summaries to Excel
    save_combined_to_excel(all_summaries, 'reports/output/summary_metrics_combined_90_days.xlsx')
    print("Data successfully saved to 'summary_metrics_combined_90_days.xlsx'")