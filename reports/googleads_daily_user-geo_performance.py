# file name:
# version:
# Notes:

import pandas as pd
import os
from google.ads.googleads.client import GoogleAdsClient
from google.ads.googleads.errors import GoogleAdsException
from datetime import datetime, timedelta
import re
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
import numpy as np

# Set the path to the Google Ads API configuration file
os.environ["GOOGLE_ADS_CONFIGURATION_FILE_PATH"] = "authentication/google-ads.yaml"

# Initialize the Google Ads client
client = GoogleAdsClient.load_from_storage()

def get_campaign_data(client, customer_id):
    ga_service = client.get_service("GoogleAdsService", version="v16")
    
    # Calculate date range for the past 52 weeks
    end_date = datetime.today().strftime('%Y-%m-%d')
    start_date = (datetime.today() - timedelta(weeks=13)).strftime('%Y-%m-%d')
    
    query = f"""
        SELECT
            segments.date,
            campaign.id,
            campaign.name,
            segments.device,
            segments.ad_network_type,
            segments.conversion_action,
            metrics.all_conversions,
            segments.geo_target
        FROM
            campaign
        WHERE
            segments.date BETWEEN '{start_date}' AND '{end_date}'
    """
    
    campaign_data = []
    try:
        response = ga_service.search_stream(customer_id=customer_id, query=query)
        for batch in response:
            for row in batch.results:
                campaign_data.append({
                    "Date": row.segments.date,
                    "Campaign ID": row.campaign.id,
                    "Campaign Name": row.campaign.name,
                    "Device": row.segments.device.name,
                    "Network": row.segments.ad_network_type.name,
                    "Conversion Action ID": row.segments.conversion_action,
                    "Conversions": row.metrics.all_conversions,
                    "Geo Target State": row.segments.geo_target
                })
    except GoogleAdsException as ex:
        print(f"Request failed with status {ex.error.code().name} and includes the following errors:")
        for error in ex.failure.errors:
            print(f"\tError with message {error.message}.")
            if error.location:
                for field_path_element in error.location.field_path_elements:
                    print(f"\t\tOn field: {field_path_element.field_name}")
        return None

    # Convert dataset to DataFrame
    campaign_df = pd.DataFrame(campaign_data)
    
    return campaign_df

def get_conversion_action_data(client, customer_id):
    ga_service = client.get_service("GoogleAdsService", version="v16")
    
    query = """
        SELECT
            conversion_action.id,
            conversion_action.name
        FROM
            conversion_action
    """
    
    conversion_action_data = []
    try:
        response = ga_service.search_stream(customer_id=customer_id, query=query)
        for batch in response:
            for row in batch.results:
                conversion_action_data.append({
                    "Conversion Action ID": row.conversion_action.id,
                    "Conversion Action Name": row.conversion_action.name
                })
    except GoogleAdsException as ex:
        print(f"Request failed with status {ex.error.code().name} and includes the following errors:")
        for error in ex.failure.errors:
            print(f"\tError with message {error.message}.")
            if error.location:
                for field_path_element in error.location.field_path_elements:
                    print(f"\t\tOn field: {field_path_element.field_name}")
        return None

    # Convert dataset to DataFrame
    conversion_action_df = pd.DataFrame(conversion_action_data)
    
    return conversion_action_df

def get_clicks_impressions_spend_data(client, customer_id):
    ga_service = client.get_service("GoogleAdsService", version="v16")
    
    # Calculate date range for the past 52 weeks
    end_date = datetime.today().strftime('%Y-%m-%d')
    start_date = (datetime.today() - timedelta(weeks=13)).strftime('%Y-%m-%d')
    
    query = f"""
        SELECT
        segments.date,
        ad_group.id,
        ad_group.name,
        campaign.id,
        campaign.name,
        metrics.clicks,
        metrics.impressions,
        metrics.cost_micros,
        segments.geo_target_state
    FROM
        ad_group_criterion
        WHERE
            segments.date BETWEEN '{start_date}' AND '{end_date}'
    """
    
    geo_data = []
    try:
        response = ga_service.search_stream(customer_id=customer_id, query=query)
        for batch in response:
            for row in batch.results:
                geo_data.append({
                    "Date": row.segments.date,
                    "Ad Group ID": row.ad_group.id,
                    "Ad Group Name": row.ad_group.name,
                    "Campaign ID": row.campaign.id,
                    "Campaign Name": row.campaign.name,
                    "Clicks": row.metrics.clicks,
                    "Impressions": row.metrics.impressions,
                    "Cost": row.metrics.cost_micros / 1e6,  # Convert micros to currency
                    "State": row.segments.geo_target_state
                })
    except GoogleAdsException as ex:
        print(f"Request failed with status {ex.error.code().name} and includes the following errors:")
        for error in ex.failure.errors:
            print(f"\tError with message {error.message}.")
            if error.location:
                for field_path_element in error.location.field_path_elements:
                    print(f"\t\tOn field: {field_path_element.field_name}")
        return None

    # Convert dataset to DataFrame
    geo_df = pd.DataFrame(geo_data)
    
    return geo_df

def extract_conversion_action_id(conversion_action_str):
    match = re.search(r'conversionActions/(\d+)', conversion_action_str)
    return match.group(1) if match else None

def save_to_excel(dataframe, filename, sheet_name='Sheet1'):
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        dataframe.to_excel(writer, sheet_name=sheet_name, index=False)

def save_combined_to_excel(data_dict, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet"
    
    sheet_names = []
    for date, data in data_dict.items():
        sheet_names.append(date)
        
        # Create a new sheet for each date
        new_ws = wb.create_sheet(title=date)
        
        row_start = 1
        for title, data in data.items():
            # Add title
            new_ws.cell(row=row_start, column=1, value=title).font = Font(size=14, bold=True)
            
            # Convert DataFrame to rows and append to worksheet
            for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), start=row_start + 1):
                for c_idx, value in enumerate(row, start=1):
                    # Ensure the value is scalar
                    if isinstance(value, np.ndarray):
                        value = value.item()  # Convert numpy array to scalar
                    
                    new_ws.cell(row=r_idx, column=c_idx, value=value)
            
            # Move to next start row after the table, plus some space
            row_start = new_ws.max_row + 3
    
    # Populate the "Sheet" worksheet with the names of the other worksheets
    for idx, sheet_name in enumerate(sheet_names, start=2):  # Start from cell B2
        ws.cell(row=2, column=idx, value=sheet_name)
    
    wb.save(filename)
    


# Replace with your actual customer ID
CUSTOMER_ID = '7554573980' # TNH: '7731032510' ; FF: '7554573980'

# Define tenant campaign names
tenant_campaigns = ["FF - Search - Brand - Broad", "Search - Tenants"]

# # Get campaign data
# campaign_data = get_campaign_data(client, CUSTOMER_ID)

# # Get conversion action data
# conversion_action_data = get_conversion_action_data(client, CUSTOMER_ID)


# Get clicks, impressions, and spend data
clicks_impressions_spend_data = get_clicks_impressions_spend_data(client, CUSTOMER_ID)

# if campaign_data is not None and conversion_action_data is not None:
#     # Extract numerical Conversion Action ID from the string format in campaign data
#     campaign_data["Conversion Action ID"] = campaign_data["Conversion Action ID"].apply(extract_conversion_action_id)
    
#     # Ensure both columns are of the same type (string in this case)
#     campaign_data["Conversion Action ID"] = campaign_data["Conversion Action ID"].astype(str)
#     conversion_action_data["Conversion Action ID"] = conversion_action_data["Conversion Action ID"].astype(str)
    
#     # Merge the two dataframes on Conversion Action ID
#     merged_data = pd.merge(campaign_data, conversion_action_data, on="Conversion Action ID", how="left")

#     # Pivot the data to have individual columns for each conversion action name
#     pivot_data = merged_data.pivot_table(index=["Date", "Campaign ID", "Device", "Network"],
#                                          columns="Conversion Action Name",
#                                          values="Conversions",
#                                          aggfunc="sum",
#                                          fill_value=0).reset_index()
    
#     # Flatten the column headers
#     pivot_data.columns.name = None

# else:
#     print("Failed to retrieve data.")

# if clicks_impressions_spend_data is not None:
#     # Merge the pivoted conversion data with the clicks, impressions, and spend data
#     final_data = pd.merge(pivot_data, clicks_impressions_spend_data, on=["Date", "Campaign ID", "Device", "Network"], how="right")

#     # Filter out campaigns containing "tenant" or "display"
#     # final_data = final_data[~final_data['Campaign Name'].str.contains('Tenants', case=False)]
#     final_data = final_data[~final_data['Campaign Name'].str.contains('Display', case=False)]
    
#     # Include only "SEARCH" or "SEARCH_PARTNERS" networks
#     final_data = final_data[final_data['Network'].str.contains('SEARCH|SEARCH_PARTNERS', case=False)]

#     # Save final data to Excel
#     save_to_excel(final_data, 'reports/output/final-geo_data.xlsx') #tnhfinal_data.xlsx'
#     print("Data successfully saved to 'final-geo_data.xlsx'")