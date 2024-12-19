# file name:
# version: V000-000-005
# output:
# Notes: Adds alternating row Fill Color and widens column a and b with padding and applies font styling & size

import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, NamedStyle, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Load the CSV into a DataFrame
df = pd.read_csv('ga4-reports/output/ga4_summary_metrics_iso_weeks_sorted.csv')

# Map for renaming columns
rename_map = {
    'Cost_Actual': 'Spend_Actual',
    'Cost_LW': 'Spend_LW',
    'Cost_YoY': 'Spend_YoY',
    'Cost_WoW': 'Spend_WoW',
    'FF_Lead_Event_Count_Actual': 'Leads_Actual',
    'FF_Lead_Event_Count_LW': 'Leads_LW',
    'FF_Lead_Event_Count_YoY': 'Leads_YoY',
    'FF_Lead_Event_Count_WoW': 'Leads_WoW',
    'FF_Purchase_Event_Count_Actual': 'New Properties_Actual',
    'FF_Purchase_Event_Count_LW': 'New Properties_LW',
    'FF_Purchase_Event_Count_YoY': 'New Properties_YoY',
    'FF_Purchase_Event_Count_WoW': 'New Properties_WoW',
    'Lead_Conversion_Actual': 'Lead Conversion_Actual',
    'Lead_Conversion_LW': 'Lead Conversion_LW',
    'Lead_Conversion_YoY': 'Lead Conversion_YoY',
    'Lead_Conversion_WoW': 'Lead Conversion_WoW',
    'FF_BRSubmit_Event_Count_Actual': 'Booking Requests_Actual',
    'FF_BRSubmit_Event_Count_LW': 'Booking Requests_LW',
    'FF_BRSubmit_Event_Count_YoY': 'Booking Requests_YoY',
    'FF_BRSubmit_Event_Count_WoW': 'Booking Requests_WoW',
    'FF_DMSubmit_Event_Count_Actual': 'Direct Messages_Actual',
    'FF_DMSubmit_Event_Count_LW': 'Direct Messages_LW',
    'FF_DMSubmit_Event_Count_YoY': 'Direct Messages_YoY',
    'FF_DMSubmit_Event_Count_WoW': 'Direct Messages_WoW',
    'FF_PhoneGet_Event_Count_Actual': 'Phone Reveals_Actual',
    'FF_PhoneGet_Event_Count_LW': 'Phone Reveals_LW',
    'FF_PhoneGet_Event_Count_YoY': 'Phone Reveals_YoY',
    'FF_PhoneGet_Event_Count_WoW': 'Phone Reveals_WoW',
    'Traveler_Conversion_Actual': 'Traveler Conversion_Actual',
    'Traveler_Conversion_LW': 'Traveler Conversion_LW',
    'Traveler_Conversion_YoY': 'Traveler Conversion_YoY',
    'Traveler_Conversion_WoW': 'Traveler Conversion_WoW'
}

# Rename columns
df.rename(columns=rename_map, inplace=True)

# Define the desired column order
columns_order = [
    'Spend_Actual', 'Spend_LW', 'Spend_WoW', 'Spend_YoY',
    'Leads_Actual', 'Leads_LW', 'Leads_WoW', 'Leads_YoY', 
    'CPL_Actual', 'CPL_LW', 'CPL_WoW', 'CPL_YoY',
    'New Properties_Actual', 'New Properties_LW', 'New Properties_WoW', 'New Properties_YoY',
    'Lead Conversion_Actual', 'Lead Conversion_LW', 'Lead Conversion_WoW', 'Lead Conversion_YoY',
    'CAC_Actual', 'CAC_LW', 'CAC_WoW', 'CAC_YoY',
    'Booking Requests_Actual', 'Booking Requests_LW', 'Booking Requests_WoW', 'Booking Requests_YoY',
    'Direct Messages_Actual', 'Direct Messages_LW', 'Direct Messages_WoW', 'Direct Messages_YoY',
    'Phone Reveals_Actual', 'Phone Reveals_LW', 'Phone Reveals_WoW', 'Phone Reveals_YoY',
    'Housing Requests_Actual', 'Housing Requests_LW', 'Housing Requests_WoW', 'Housing Requests_YoY',
    'Traveler Conversion_Actual', 'Traveler Conversion_LW', 'Traveler Conversion_WoW', 'Traveler Conversion_YoY',
    'CPTA_Actual', 'CPTA_LW', 'CPTA_WoW', 'CPTA_YoY',
    'Traveler Value_Actual', 'Traveler Value_LW', 'Traveler Value_WoW', 'Traveler Value_YoY',
    'Landlord Value_Actual', 'Landlord Value_LW', 'Landlord Value_WoW', 'Landlord Value_YoY',
    'ROAS_Actual', 'ROAS_LW', 'ROAS_WoW', 'ROAS_YoY',
    'Impressions_Actual', 'Impressions_LW', 'Impressions_WoW', 'Impressions_YoY',
    'Clicks_Actual', 'Clicks_LW', 'Clicks_WoW', 'Clicks_YoY',
    'Sessions_Actual', 'Sessions_LW', 'Sessions_WoW', 'Sessions_YoY'
]

# Verify that the desired columns exist in the DataFrame
df = df[['Channel_Group', 'Campaign_Group'] + [col for col in columns_order if col in df.columns]]

# Save to a new Excel workbook
wb = openpyxl.Workbook()
ws = wb.active

# Insert empty rows at the top
for _ in range(4):
    ws.append([])

# Add the DataFrame to the worksheet
for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

# Set the base font style and size across the entire worksheet
base_font = Font(name="Aptos Display", size=12)

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.font = base_font

# Format the numeric columns before splitting the metric names
dollar_no_decimal = NamedStyle(name="dollar_no_decimal", font=base_font)
dollar_no_decimal.number_format = '"$"#,##0'

dollar_two_decimal = NamedStyle(name="dollar_two_decimal", font=base_font)
dollar_two_decimal.number_format = '"$"#,##0.00'

percent_no_decimal = NamedStyle(name="percent_no_decimal", font=base_font)
percent_no_decimal.number_format = '0%'

percent_two_decimal = NamedStyle(name="percent_two_decimal", font=base_font)
percent_two_decimal.number_format = '0.00%'

comma_no_decimal = NamedStyle(name="comma_no_decimal", font=base_font)
comma_no_decimal.number_format = '#,##0'

# Apply formats based on the column names
for col in ws.iter_cols(min_col=3, max_col=ws.max_column, min_row=5, max_row=ws.max_row):
    header = col[0].value
    for cell in col[1:]:
        if header.endswith('Actual') or header.endswith('LW'):
            if 'Spend' in header or 'Traveler Value' in header or 'Landlord Value' in header:
                cell.style = dollar_no_decimal
            elif 'CPL' in header or 'CAC' in header:
                cell.style = dollar_no_decimal
            elif 'CPTA' in header:
                cell.style = dollar_two_decimal
            elif 'ROAS' in header:
                cell.style = percent_no_decimal
            elif 'Lead Conversion' in header or 'Traveler Conversion' in header:
                cell.style = percent_two_decimal
            else:
                cell.style = comma_no_decimal
        elif header.endswith('WoW') or header.endswith('YoY'):
            cell.style = percent_no_decimal

# Now proceed to split the metric names into two rows
# Create a list of the columns to format
metric_names = [col.split("_")[0] for col in columns_order]
unique_metric_names = sorted(set(metric_names), key=metric_names.index)

# Insert a new row above the current column headers for the metric names
ws.insert_rows(5)

# Adjusted column offset to 2 to account for Channel_Group and Campaign_Group columns
col_offset = 2  # Offset to account for Channel_Group and Campaign_Group columns
for idx, metric_name in enumerate(unique_metric_names):
    start_col = idx * 4 + col_offset + 1
    end_col = start_col + 3

    # Set the metric name in the new row
    ws.cell(row=5, column=start_col, value=metric_name)

    # Merge and center the metric name cells over the four periods
    ws.merge_cells(start_row=5, start_column=start_col, end_row=5, end_column=end_col)
    ws.cell(row=5, column=start_col).alignment = Alignment(horizontal="center")

# Adjust the original header row to show only the period names
for col in ws.iter_cols(min_col=col_offset + 1, max_col=ws.max_column, min_row=6, max_row=6):
    period_label = col[0].value.split("_")[-1]
    col[0].value = period_label

# Define the border style for the outside border
thin_side = Side(style='thin')

# Apply outside borders around each metric column group
for idx, metric_name in enumerate(unique_metric_names):
    start_col = idx * 4 + col_offset + 1
    end_col = start_col + 3
    
    # Apply borders from the top row (5) down to the last row in the data
    for row in range(5, ws.max_row + 1):
        # Set left and right borders for each cell in the group
        ws.cell(row=row, column=start_col).border = Border(left=thin_side)
        ws.cell(row=row, column=end_col).border = Border(right=thin_side)
        ws.cell(row=row, column=start_col).font = base_font  # Ensure font consistency
        ws.cell(row=row, column=end_col).font = base_font
    
    # Apply top and bottom borders and ensure side borders are maintained
    for col in range(start_col, end_col + 1):
        ws.cell(row=5, column=col).border = Border(top=thin_side, left=thin_side if col == start_col else None, right=thin_side if col == end_col else None)
        ws.cell(row=ws.max_row, column=col).border = Border(bottom=thin_side, left=thin_side if col == start_col else None, right=thin_side if col == end_col else None)
        ws.cell(row=5, column=col).font = base_font  # Ensure font consistency
        ws.cell(row=ws.max_row, column=col).font = base_font


# 1. Apply a white "Fill Color" to the entire workbook
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

for row in ws.iter_rows(min_row=1, max_row=ws.max_row +20, min_col=1, max_col=ws.max_column +20):
    for cell in row:
        cell.fill = white_fill

# 2. Apply a "Fill Color" of hex #024991 to row 6 (headers) from column C to BV
header_fill = PatternFill(start_color="024991", end_color="024991", fill_type="solid")
header_font = Font(color="FFFFFF", name="Aptos Display", size=12)

for col in ws.iter_cols(min_col=1, max_col=74, min_row=6, max_row=6):  # C = 3, BV = 70
    for cell in col:
        cell.fill = header_fill
        # 3. Change the font color to white in the same cells
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')


# 1. Apply alternating row colors starting from the first data row to the last row and column of data
alt_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

# Start alternating fill from the first row with data (row 7 in your case)
for row in range(7, ws.max_row + 1):
    if (row - 7) % 2 == 0:  # Alternate every other row
        for col in range(1, ws.max_column + 1):  # Apply to each column in the range
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:  # Only apply if the cell has data
                cell.fill = alt_fill

# 2. Adjust column width for column A & B
for col in ['A', 'B']:
    max_length = 0
    column = ws[col]
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2  # Adjust width with some padding
    ws.column_dimensions[col].width = adjusted_width

# ----------------- Channel Group Bolding & row 5 Fill Colors ------------------------ #

# Apply bold font to "Channel Group" and "Campaign Group" columns for rows containing "SEM" or "Paid"
bold_font = Font(bold=True, name="Aptos Display", size=12)

for row in range(7, ws.max_row + 1):  # Start from row 7 where data begins
    channel_group_cell = ws.cell(row=row, column=1)  # "Channel Group" is in column A (1)
    campaign_group_cell = ws.cell(row=row, column=2)  # "Campaign Group" is in column B (2)
    
    # Check if "SEM" or "Paid" is in the Channel Group or Campaign Group
    if "SEM" in str(channel_group_cell.value) or "Paid" in str(channel_group_cell.value):
        channel_group_cell.font = bold_font
        
    if "SEM" in str(campaign_group_cell.value) or "Paid" in str(campaign_group_cell.value):
        campaign_group_cell.font = bold_font


# Define the fill colors and bold font
fill_color_1 = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
fill_color_2 = PatternFill(start_color="CAEDFB", end_color="CAEDFB", fill_type="solid")
bold_font = Font(bold=True, name="Aptos Display", size=12)

# Iterate through the metric names and apply the styles
for idx, metric_name in enumerate(unique_metric_names):
    start_col = idx * 4 + col_offset + 1
    end_col = start_col + 3
    
    # Alternate between the two fill colors
    if idx % 2 == 0:
        fill_color = fill_color_1
    else:
        fill_color = fill_color_2
    
    # Apply the styles to the merged cells
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=5, column=col)
        cell.font = bold_font
        cell.fill = fill_color

# Freeze columns A & B so that they stay visible while scrolling to the right
ws.freeze_panes = 'C1'

# Save the workbook
wb.save('ga4-reports/output/ga4_summary_metrics_formatted.xlsx')

print("Formatted Excel file with borders has been created and saved.")
