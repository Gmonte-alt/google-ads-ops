# file name: ga4-reports/excel_workbook_summary_ga4_ad-platform_wtd.py
# version: V000-000-021
# input: ga4-reports/input/xChannelGrowthMarketing_Read_Me.xlsx 'ga4-reports/input/Picture1.png
# Output: Growth-Marketing-X-Channel-Performance-report.xlsx
# Notes: Integrates with xlsxwriter chart file created in new previous step

import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, NamedStyle, PatternFill, Border, Side
from openpyxl import load_workbook
import gspread
from google.oauth2.service_account import Credentials
from datetime import datetime



# -------- Google Sheets Integration -------- #

# Path to your service account JSON key file
SERVICE_ACCOUNT_FILE = 

# Define the scopes required for the Google Sheets API
SCOPES = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive']

# Create a credentials object using the service account file
creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)

# Authenticate and initialize the gspread client
client = gspread.authorize(creds)

# Open the Google Sheet by its name or URL
sheet = client.open('FF - Growth - Weekly Commentary').sheet1

# Get all rows from the Google Sheet
all_rows = sheet.get_all_values()
headers = all_rows[0]  # First row as headers
data = all_rows[1:]  # Data rows

# -------- Excel Workbook Creation and Formatting -------- #

excel_file_path = 'ga4-reports/output/ga4_summary_metrics_formatted.xlsx'

# Load the CSV into a DataFrame
df = pd.read_csv('ga4-reports/output/ga4_summary_metrics_multiple_periods.csv')

# List of periods
periods = ['WTD', 'EOW', 'MTD', 'EOM', 'QTD', 'EOQ', 'YTD']

# Create a new Excel workbook
wb = openpyxl.Workbook()

# Set the base font style and size across the entire worksheet
base_font = Font(name="Aptos Display", size=12)

# Define or get existing formatting styles
dollar_no_decimal = NamedStyle(name="dollar_no_decimal", font=base_font)
dollar_no_decimal.number_format = '"$"#,##0'
wb.add_named_style(dollar_no_decimal)

dollar_two_decimal = NamedStyle(name="dollar_two_decimal", font=base_font)
dollar_two_decimal.number_format = '"$"#,##0.00'
wb.add_named_style(dollar_two_decimal)

percent_no_decimal = NamedStyle(name="percent_no_decimal", font=base_font)
percent_no_decimal.number_format = '0%'
wb.add_named_style(percent_no_decimal)

percent_two_decimal = NamedStyle(name="percent_two_decimal", font=base_font)
percent_two_decimal.number_format = '0.00%'
wb.add_named_style(percent_two_decimal)

comma_no_decimal = NamedStyle(name="comma_no_decimal", font=base_font)
comma_no_decimal.number_format = '#,##0'
wb.add_named_style(comma_no_decimal)


# Iterate over each period
for period in periods:
    # Filter the DataFrame for the current period
    period_df = df[df['Period'] == period]

    # Determine the start and end dates
    start_date = period_df['Start_Date_Period'].min()
    end_date = period_df['End_Date_Period'].max()

    # Drop the unnecessary columns
    period_df = period_df.drop(columns=['Start_Date_Period', 'End_Date_Period', 'Period'])

    # Rename columns according to the mapping
    rename_map = {
        'FF_Lead_Event_Count': 'Leads_Actual',
        'FF_Purchase_Event_Count': 'New Properties_Actual',
        'FF_BRSubmit_Event_Count': 'Booking Requests_Actual',
        'FF_DMSubmit_Event_Count': 'Direct Messages_Actual',
        'FF_PhoneGet_Event_Count': 'Phone Reveals_Actual',
        'Housing Requests': 'Housing Requests_Actual',
        'Total Traveler Actions': 'Total Traveler Actions_Actual',
        'Traveler Value': 'Traveler Value_Actual',
        'Landlord Value': 'Landlord Value_Actual',
        'Sessions': 'Sessions_Actual',
        'Impressions': 'Impressions_Actual',
        'Clicks': 'Clicks_Actual',
        'Cost': 'Spend_Actual',
        'Lead_Conversion': 'Lead Conversion_Actual',
        'CPL': 'CPL_Actual',
        'CAC': 'CAC_Actual',
        'Traveler_Conversion': 'Traveler Conversion_Actual',
        'CPTA': 'CPTA_Actual',
        'ROAS': 'ROAS_Actual',
        'CPC': 'CPC_Actual',  # New Metric: Cost Per Click
        'CTR': 'CTR_Actual',  # New Metric: Click Through Rate
        'RPC_Landlord': 'RPC Landlord_Actual',  # New Metric: Revenue Per Click - Landlord
        'RPC_Tenant': 'RPC Tenant_Actual',  # New Metric: Revenue Per Click - Tenant
        'Purchase_to_Sessions': 'Purchase to Sessions CVR_Actual', # New Metrics: Purchase / Sessions CVR
        'Lead_to_Sessions': 'Lead to Sessions CVR_Actual', # New Metrics: Lead / Sessions CVR
        'Impression_Share': 'Impression Share_Actual', # New Impression Share

        'FF_Lead_Event_Count_previous': 'Leads_LW',
        'FF_Purchase_Event_Count_previous': 'New Properties_LW',
        'FF_BRSubmit_Event_Count_previous': 'Booking Requests_LW',
        'FF_DMSubmit_Event_Count_previous': 'Direct Messages_LW',
        'FF_PhoneGet_Event_Count_previous': 'Phone Reveals_LW',
        'Housing Requests_previous': 'Housing Requests_LW',
        'Total Traveler Actions_previous': 'Total Traveler Actions_LW',
        'Traveler Value_previous': 'Traveler Value_LW',
        'Landlord Value_previous': 'Landlord Value_LW',
        'Sessions_previous': 'Sessions_LW',
        'Impressions_previous': 'Impressions_LW',
        'Clicks_previous': 'Clicks_LW',
        'Cost_previous': 'Spend_LW',
        'Lead_Conversion_previous': 'Lead Conversion_LW',
        'CPL_previous': 'CPL_LW',
        'CAC_previous': 'CAC_LW',
        'Traveler_Conversion_previous': 'Traveler Conversion_LW',
        'CPTA_previous': 'CPTA_LW',
        'ROAS_previous': 'ROAS_LW',
        'CPC_previous': 'CPC_LW',  # New Metric: CPC Last Week
        'CTR_previous': 'CTR_LW',  # New Metric: CTR Last Week
        'RPC_Landlord_previous': 'RPC Landlord_LW',  # New Metric: RPC Landlord Last Week
        'RPC_Tenant_previous': 'RPC Tenant_LW',  # New Metric: RPC Tenant Last Week
        'Purchase_to_Sessions_previous': 'Purchase to Sessions CVR_LW', # New Metrics: Purchase / Sessions CVR
        'Lead_to_Sessions_previous': 'Lead to Sessions CVR_LW', # New Metrics: Lead / Sessions CVR
        'Impression_Share_previous': 'Impression Share_LW', # New Impression Share

        'FF_Lead_Event_Count_WoW': 'Leads_WoW',
        'FF_Purchase_Event_Count_WoW': 'New Properties_WoW',
        'FF_BRSubmit_Event_Count_WoW': 'Booking Requests_WoW',
        'FF_DMSubmit_Event_Count_WoW': 'Direct Messages_WoW',
        'FF_PhoneGet_Event_Count_WoW': 'Phone Reveals_WoW',
        'Housing Requests_WoW': 'Housing Requests_WoW',
        'Total Traveler Actions_WoW': 'Total Traveler Actions_WoW',
        'Traveler Value_WoW': 'Traveler Value_WoW',
        'Landlord Value_WoW': 'Landlord Value_WoW',
        'Sessions_WoW': 'Sessions_WoW',
        'Impressions_WoW': 'Impressions_WoW',
        'Clicks_WoW': 'Clicks_WoW',
        'Cost_WoW': 'Spend_WoW',
        'Lead_Conversion_WoW': 'Lead Conversion_WoW',
        'CPL_WoW': 'CPL_WoW',
        'CAC_WoW': 'CAC_WoW',
        'Traveler_Conversion_WoW': 'Traveler Conversion_WoW',
        'CPTA_WoW': 'CPTA_WoW',
        'ROAS_WoW': 'ROAS_WoW',
        'CPC_WoW': 'CPC_WoW',  # New Metric: CPC WoW
        'CTR_WoW': 'CTR_WoW',  # New Metric: CTR WoW
        'RPC_Landlord_WoW': 'RPC Landlord_WoW',  # New Metric: RPC Landlord WoW
        'RPC_Tenant_WoW': 'RPC Tenant_WoW',  # New Metric: RPC Tenant WoW
        'Purchase_to_Sessions_WoW': 'Purchase to Sessions CVR_WoW', # New Metrics: Purchase / Sessions CVR
        'Lead_to_Sessions_WoW': 'Lead to Sessions CVR_WoW', # New Metrics: Lead / Sessions CVR
        'Impression_Share_WoW': 'Impression Share_WoW', # New Metrics: Impression Share

        'FF_Lead_Event_Count_YoY': 'Leads_YoY',
        'FF_Purchase_Event_Count_YoY': 'New Properties_YoY',
        'FF_BRSubmit_Event_Count_YoY': 'Booking Requests_YoY',
        'FF_DMSubmit_Event_Count_YoY': 'Direct Messages_YoY',
        'FF_PhoneGet_Event_Count_YoY': 'Phone Reveals_YoY',
        'Housing Requests_YoY': 'Housing Requests_YoY',
        'Total Traveler Actions_YoY': 'Total Traveler Actions_YoY',
        'Traveler Value_YoY': 'Traveler Value_YoY',
        'Landlord Value_YoY': 'Landlord Value_YoY',
        'Sessions_YoY': 'Sessions_YoY',
        'Impressions_YoY': 'Impressions_YoY',
        'Clicks_YoY': 'Clicks_YoY',
        'Cost_YoY': 'Spend_YoY',
        'Lead_Conversion_YoY': 'Lead Conversion_YoY',
        'CPL_YoY': 'CPL_YoY',
        'CAC_YoY': 'CAC_YoY',
        'Traveler_Conversion_YoY': 'Traveler Conversion_YoY',
        'CPTA_YoY': 'CPTA_YoY',
        'ROAS_YoY': 'ROAS_YoY',
        'CPC_YoY': 'CPC_YoY',  # New Metric: CPC YoY
        'CTR_YoY': 'CTR_YoY',  # New Metric: CTR YoY
        'RPC_Landlord_YoY': 'RPC Landlord_YoY',  # New Metric: RPC Landlord YoY
        'RPC_Tenant_YoY': 'RPC Tenant_YoY',  # New Metric: RPC Tenant YoY
        'Purchase_to_Sessions_YoY': 'Purchase to Sessions CVR_YoY', # New Metrics: Purchase / Sessions CVR
        'Lead_to_Sessions_YoY': 'Lead to Sessions CVR_YoY', # New Metrics: Lead / Sessions CVR
        'Impression_Share_YoY': 'Impression Share_YoY', # New Metrics: Impression Share
    }


    period_df.rename(columns=rename_map, inplace=True)

    # Define the desired column order
    columns_order = [
        'Spend_Actual', 'Spend_LW', 'Spend_WoW', 'Spend_YoY',
        'New Properties_Actual', 'New Properties_LW', 'New Properties_WoW', 'New Properties_YoY',
        'CAC_Actual', 'CAC_LW', 'CAC_WoW', 'CAC_YoY',
        'Total Traveler Actions_Actual', 'Total Traveler Actions_LW', 'Total Traveler Actions_WoW', 'Total Traveler Actions_YoY',
        'Traveler Conversion_Actual', 'Traveler Conversion_LW', 'Traveler Conversion_WoW', 'Traveler Conversion_YoY',
        'CPTA_Actual', 'CPTA_LW', 'CPTA_WoW', 'CPTA_YoY',
        'ROAS_Actual', 'ROAS_LW', 'ROAS_WoW', 'ROAS_YoY',
        'Traveler Value_Actual', 'Traveler Value_LW', 'Traveler Value_WoW', 'Traveler Value_YoY',
        'Landlord Value_Actual', 'Landlord Value_LW', 'Landlord Value_WoW', 'Landlord Value_YoY',
        'Booking Requests_Actual', 'Booking Requests_LW', 'Booking Requests_WoW', 'Booking Requests_YoY',
        'Direct Messages_Actual', 'Direct Messages_LW', 'Direct Messages_WoW', 'Direct Messages_YoY',
        'Phone Reveals_Actual', 'Phone Reveals_LW', 'Phone Reveals_WoW', 'Phone Reveals_YoY',
        'Housing Requests_Actual', 'Housing Requests_LW', 'Housing Requests_WoW', 'Housing Requests_YoY',
        'Leads_Actual', 'Leads_LW', 'Leads_WoW', 'Leads_YoY',
        'Lead Conversion_Actual', 'Lead Conversion_LW', 'Lead Conversion_WoW', 'Lead Conversion_YoY',
        'CPL_Actual', 'CPL_LW', 'CPL_WoW', 'CPL_YoY',
        
        # New Metrics: CPC, CTR, RPC Landlord, RPC Tenant
        'CPC_Actual', 'CPC_LW', 'CPC_WoW', 'CPC_YoY',  # Cost per Click
        'RPC Landlord_Actual', 'RPC Landlord_LW', 'RPC Landlord_WoW', 'RPC Landlord_YoY',  # Revenue per Click - Landlord
        'RPC Tenant_Actual', 'RPC Tenant_LW', 'RPC Tenant_WoW', 'RPC Tenant_YoY',  # Revenue per Click - Tenant
        
        'Impressions_Actual', 'Impressions_LW', 'Impressions_WoW', 'Impressions_YoY',
        'Clicks_Actual', 'Clicks_LW', 'Clicks_WoW', 'Clicks_YoY',
        'Sessions_Actual', 'Sessions_LW', 'Sessions_WoW', 'Sessions_YoY',
        'CTR_Actual', 'CTR_LW', 'CTR_WoW', 'CTR_YoY'  # Click Through Rate
        'Purchase to Sessions CVR_Actual', 'Purchase to Sessions CVR_LW', 'Purchase to Sessions CVR_WoW', 'Purchase to Sessions CVR_YoY',
        'Lead to Sessions CVR_Actual', 'Lead to Sessions CVR_LW', 'Lead to Sessions CVR_WoW', 'Lead to Sessions CVR_YoY',
        'Impression Share_Actual', 'Impression Share_LW', 'Impression Share_WoW', 'Impression Share_YoY'
    ]


    # Reorder the columns according to the specified order
    period_df = period_df[['Channel_Group', 'Campaign_Group'] + [col for col in columns_order if col in period_df.columns]]

    # Add a new worksheet for the period
    ws = wb.create_sheet(title=period)

    # Set the title and sub-title
    ws['A1'] = f"GROWTH - PERFORMANCE REPORT - {period}"
    ws['A1'].font = Font(name="Aptos Display", size=18, bold=True, color="FFFFFF")
    ws['A1'].alignment = Alignment(horizontal="left")

    ws['A2'] = f"{start_date} to {end_date}"
    ws['A2'].font = Font(name="Aptos Display", size=16, bold=True, color="FFFFFF")
    ws['A2'].alignment = Alignment(horizontal="left")

    # Insert empty rows for padding between header and table
    ws.append([])  # Insert an empty row
    ws.append([])  # Insert another empty row

    # Insert the DataFrame into the worksheet
    for r in dataframe_to_rows(period_df, index=False, header=True):
        ws.append(r)
    
    # ------------- Formats of metrics based on column names --------------- #
    
    # Apply formats based on the column names
    for col in ws.iter_cols(min_col=3, max_col=ws.max_column, min_row=5, max_row=ws.max_row):
        header = col[0].value
        if header is not None:  # Ensure header is not None before processing
            print(f"Formatting column: {header}")  # Debugging line to check header
            for cell in col[1:]:
                if header.endswith('Actual') or header.endswith('LW'):
                    if 'Spend' in header or 'Traveler Value' in header or 'Landlord Value' in header:
                        cell.style = dollar_no_decimal
                    elif 'CPL' in header or 'CAC' in header:
                        cell.style = dollar_no_decimal
                    elif 'CPTA' in header or 'CPC' in header or 'RPC Landlord' in header or 'RPC Tenant' in header:
                        cell.style = dollar_two_decimal
                    elif 'ROAS' in header:
                        cell.style = percent_no_decimal
                    elif 'Lead Conversion' in header or 'Traveler Conversion' in header:
                        cell.style = percent_no_decimal
                    elif 'CTR' in header or 'Purchase to Sessions CVR' in header or 'Lead to Sessions CVR' in header:
                        cell.style = percent_two_decimal
                    else:
                        cell.style = comma_no_decimal
                elif header.endswith('WoW') or header.endswith('YoY'):
                    cell.style = percent_no_decimal
        
    # Insert the new row and split metric names
    # Now proceed to split the metric names into two rows
    # Create a list of the columns to format
    metric_names = [col.split("_")[0] for col in columns_order]
    unique_metric_names = sorted(set(metric_names), key=metric_names.index)

    # Insert a new row above the current column headers for the metric names
    ws.insert_rows(5)

    # Adjusted column offset to 2 to account for Channel_Group and Campaign_Group columns
    col_offset = 2  # Offset to account for Channel_Group and Campaign_Group columns
    for idx, metric_name in enumerate(unique_metric_names):
        start_col = idx * 4 + col_offset + 1
        end_col = start_col + 3

        # Set the metric name in the new row
        ws.cell(row=5, column=start_col, value=metric_name)

        # Merge and center the metric name cells over the four periods
        ws.merge_cells(start_row=5, start_column=start_col, end_row=5, end_column=end_col)
        ws.cell(row=5, column=start_col).alignment = Alignment(horizontal="center")

    # Adjust the original header row to show only the period names
    for col in ws.iter_cols(min_col=col_offset + 1, max_col=ws.max_column, min_row=6, max_row=6):
        period_label = col[0].value
        if period_label is not None:  # Ensure period_label is not None before processing
            col[0].value = period_label.split("_")[-1]
    
    # Define the border style for the outside border
    thin_side = Side(style='thin')

    # Apply outside borders around each metric column group
    for idx, metric_name in enumerate(unique_metric_names):
        start_col = idx * 4 + col_offset + 1
        end_col = start_col + 3
        
        # Apply borders from the top row (5) down to the last row in the data
        for row in range(5, ws.max_row + 1):
            # Set left and right borders for each cell in the group
            ws.cell(row=row, column=start_col).border = Border(left=thin_side)
            ws.cell(row=row, column=end_col).border = Border(right=thin_side)
            ws.cell(row=row, column=start_col).font = base_font  # Ensure font consistency
            ws.cell(row=row, column=end_col).font = base_font
        
        # Apply top and bottom borders and ensure side borders are maintained
        for col in range(start_col, end_col + 1):
            ws.cell(row=5, column=col).border = Border(top=thin_side, left=thin_side if col == start_col else None, right=thin_side if col == end_col else None)
            ws.cell(row=ws.max_row, column=col).border = Border(bottom=thin_side, left=thin_side if col == start_col else None, right=thin_side if col == end_col else None)
            ws.cell(row=5, column=col).font = base_font  # Ensure font consistency
            ws.cell(row=ws.max_row, column=col).font = base_font
            
    bold_font = Font(bold=True, name="Aptos Display", size=12)
    
    fill_color_1 = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    fill_color_2 = PatternFill(start_color="CAEDFB", end_color="CAEDFB", fill_type="solid")
    bold_font = Font(bold=True, name="Aptos Display", size=12)

    # Iterate through the metric names and apply the styles
    for idx, metric_name in enumerate(unique_metric_names):
        start_col = idx * 4 + col_offset + 1
        end_col = start_col + 3
        
        # Alternate between the two fill colors
        if idx % 2 == 0:
            fill_color = fill_color_1
        else:
            fill_color = fill_color_2
        
        # Apply the styles to the merged cells
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=5, column=col)
            cell.font = bold_font
            cell.fill = fill_color


    for row in range(7, ws.max_row + 1):  # Start from row 7 where data begins
        channel_group_cell = ws.cell(row=row, column=1)  # "Channel Group" is in column A (1)
        campaign_group_cell = ws.cell(row=row, column=2)  # "Campaign Group" is in column B (2)
        
        # Check if "SEM" or "Paid" is in the Channel Group or Campaign Group
        if "SEM" in str(channel_group_cell.value) or "Paid" in str(channel_group_cell.value) or "Total" in str(channel_group_cell.value):
            channel_group_cell.font = bold_font
            
        if "SEM" in str(campaign_group_cell.value) or "Paid" in str(campaign_group_cell.value) or "Total" in str(campaign_group_cell.value):
            campaign_group_cell.font = bold_font
        
        # Check if "Grand Total" is in the Channel Group or Campaign Group
        if "Grand Total" in str(channel_group_cell.value):
            channel_group_cell.font = bold_font  # Apply bold font to "Grand Total" in Channel Group
            
        if "Grand Total" in str(campaign_group_cell.value):
            campaign_group_cell.font = bold_font  # Apply bold font to "Grand Total" in Campaign Group

    
    # ------------------- Worksheet Color and Font Formatting ------------------- #
            
    # Apply White Fill Color to the Entire Workbook
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row + 50, min_col=1, max_col=ws.max_column + 50):
        for cell in row:
            cell.fill = white_fill
    
    # Apply Fill Color to Headers and Change Font Color to White
    header_fill = PatternFill(start_color="024991", end_color="024991", fill_type="solid")
    header_font = Font(color="FFFFFF", name="Aptos Display", size=12)

    for col in ws.iter_cols(min_col=1, max_col=94, min_row=6, max_row=6):
        for cell in col:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
    
    # Apply Alternating Row Colors
    alt_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    
    for row in range(7, ws.max_row + 1):
        if (row - 7) % 2 == 0:
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    cell.fill = alt_fill
    
    # Adjust Column Width for Columns A & B
    for col in ['A', 'B']:
        max_length = 0
        column = ws[col]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[col].width = adjusted_width
    
    # Apply Font Styling and Fill Color to the Title in the First Row
    ws['A1'].font = Font(name="Aptos Display", size=18, bold=False, color="FFFFFF")
    ws['A1'].alignment = Alignment(horizontal="left")
    for cell in ws["1:1"]:
        cell.fill = PatternFill(start_color="024991", end_color="024991", fill_type="solid")
    
    # Apply Font Styling and Fill Color to the Subtitle in the Second Row
    ws['A2'].font = Font(name="Aptos Display", size=16, bold=False, color="FFFFFF")
    ws['A2'].alignment = Alignment(horizontal="left")
    for cell in ws["2:2"]:
        cell.fill = PatternFill(start_color="D65488", end_color="D65488", fill_type="solid")

    # ---------------------- Add Formatting Logic Here ---------------------- #
    
    # ------------------------------------------------------------------------ #

    # Adjust column widths based on content
    for col in ws.iter_cols(min_col=1, max_col=ws.max_column, min_row=1, max_row=ws.max_row):
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
        
    # Adjust the width of each metric column
    min_width = 10  # Minimum width for the columns

    for col in ws.iter_cols(min_col=3, max_col=ws.max_column, min_row=7, max_row=ws.max_row):
        max_length = 0
        column = col[0].column_letter  # Get the column letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max((max_length + 2) * 1.2, min_width)  # Adjust width with padding
        ws.column_dimensions[column].width = adjusted_width
        
    # Freeze columns A & B so that they stay visible while scrolling to the right
    ws.freeze_panes = 'C7'


# Remove the default "Sheet" created by openpyxl
del wb['Sheet']

# Save the workbook
wb.save(excel_file_path)

print("Formatted Excel file with multiple sheets has been created and saved.")


# ----------------------------------------------------------------------------------- #
# ------------------------ Imports ISO Weekdata for Trended View -------------------- #
# ----------------------------------------------------------------------------------- #

def add_wtd_data_tab_to_workbook(workbook, df_wtd_adjusted):
    # Check if '--DATA--WTD--' sheet already exists and remove it if necessary
    if '--DATA--WTD--' in workbook.sheetnames:
        del workbook['--DATA--WTD--']
    
    # Create a new worksheet for the WTD data
    ws = workbook.create_sheet(title='--DATA--WTD--')
    
    # Get unique ISO weeks
    unique_iso_weeks = df_wtd_adjusted['ISO_Week'].unique()

    # Identify all the metric columns to iterate over
    metric_columns = [col for col in df_wtd_adjusted.columns if col not in ['ISO_Year', 'ISO_Week', 'Channel_Group', 'Campaign_Group']]

    # Create column headers with an additional 'Metric' column
    header = ['Metric', 'Year', 'Channel_Group', 'Campaign_Group'] + [f'ISO_Week_{week}' for week in unique_iso_weeks]
    
    # Write the header to the first row of the worksheet
    ws.append(header)
    
    # Modify the ISO week column headings in the header row
    for col_num, col_name in enumerate(header, 1):
        if col_name.startswith('ISO_Week_'):
            iso_week_number = col_name.split('_')[-1]
            new_header = f"ISO {iso_week_number}"
            ws.cell(row=1, column=col_num).value = new_header


    # Iterate over each metric to add data to the worksheet
    for metric in metric_columns:
        # Iterate over each unique combination of Year, Channel_Group, and Campaign_Group
        for _, group in df_wtd_adjusted.groupby(['ISO_Year', 'Channel_Group', 'Campaign_Group']):
            # Create a row with the metric name and group information
            row_data = [metric, group['ISO_Year'].iloc[0], group['Channel_Group'].iloc[0], group['Campaign_Group'].iloc[0]]
            
            # Add metric values for each ISO week
            for week in unique_iso_weeks:
                value = group[group['ISO_Week'] == week][metric].sum()  # Use the current metric column
                row_data.append(value)
            
            # Write the row data to the worksheet
            ws.append(row_data)
    
    # Format the header
    for col_num, col_name in enumerate(header, 1):
        ws.cell(row=1, column=col_num).font = Font(bold=True)
        ws.cell(row=1, column=col_num).alignment = Alignment(horizontal='center')
    
    # Adjust column widths for better visibility
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_length + 2, 15)  # Minimum width is 15
    
    # Save the updated workbook
    workbook.save(excel_file_path)
    print(f"Worksheet '--DATA--WTD--' has been added to {excel_file_path}.")


# Load the existing Excel workbook
workbook = load_workbook(excel_file_path)

# Path to the WTD CSV file created in the previous application
wtd_csv_path = 'ga4-reports/output/ga4_wtd_adjusted_data.csv'

# Load the WTD data into a DataFrame
df_wtd_adjusted = pd.read_csv(wtd_csv_path)

# Call the function to add the WTD data to the workbook
add_wtd_data_tab_to_workbook(workbook, df_wtd_adjusted)



# ----------------------------------------------------------------------------------- #
# -------------------- Create Data Viz ISO Weekdata for Trended View ---------------- #
# ----------------------------------------------------------------------------------- #






# ----------------------------------------------------------------------------------- #
# -------------- Google Sheets to Excel Workbook Update and Formatting -------------- #
# ----------------------------------------------------------------------------------- #


# Function to insert commentary into a specified worksheet
def insert_commentary(worksheet_name):
    worksheet = workbook[worksheet_name]

    # Retrieve the start date from the second row, column A, in the worksheet
    date_range = worksheet['A2'].value
    start_date_str = date_range.split(' to ')[0]
    start_date = datetime.strptime(start_date_str, '%m/%d/%Y')

    # Calculate the ISO year and ISO week from the start date
    iso_year = start_date.isocalendar()[0]
    iso_week = start_date.isocalendar()[1]

    # Filter the Google Sheets data based on the calculated ISO year and ISO week
    filtered_data = [row for row in data if int(row[headers.index('Year')]) == iso_year and int(row[headers.index('ISO Week')]) == iso_week]

    # Ensure Channel Group and Campaign Group are preserved
    filtered_data = [[value for i, value in enumerate(row) if headers[i] not in ('Year', 'ISO Week')] for row in filtered_data]
    filtered_headers = [header for header in headers if header not in ('Year', 'ISO Week')]

    # Define the border style for the cells
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    # Specify the row and column where you want to start inserting the data
    start_row = 51  # Changed to compensate for the "Grand Total" row
    start_column = 1  # Column A

    # Insert the filtered headers into the Excel sheet and apply formatting
    for col_num, header in enumerate(filtered_headers, start=start_column):
        cell = worksheet.cell(row=start_row, column=col_num, value=header)
        # Apply color fill, white font, and center alignment to the header row
        cell.fill = PatternFill(start_color="024991", end_color="024991", fill_type="solid")
        cell.font = Font(color="FFFFFF", name="Aptos Display", size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Insert the new column header "Next 7 Days"
    worksheet.cell(row=start_row, column=19, value="Next 7 Days")
    next7days_header_cell = worksheet.cell(row=start_row, column=19)
    next7days_header_cell.fill = PatternFill(start_color="024991", end_color="024991", fill_type="solid")
    next7days_header_cell.font = Font(color="FFFFFF", name="Aptos Display", size=12)
    next7days_header_cell.alignment = Alignment(horizontal='center', vertical='center')
    next7days_header_cell.border = thin_border

    # Merge columns C to R in the header row for "Weekly Commentary"
    worksheet.merge_cells(start_row=start_row, start_column=3, end_row=start_row, end_column=18)
    header_cell = worksheet.cell(row=start_row, column=3)
    header_cell.fill = PatternFill(start_color="024991", end_color="024991", fill_type="solid")
    header_cell.font = Font(color="FFFFFF", name="Aptos Display", size=12)
    header_cell.alignment = Alignment(horizontal='center', vertical='center')
    header_cell.border = thin_border

    # Merge columns S to AH for the "Next 7 Days" in the header row and apply formatting
    worksheet.merge_cells(start_row=start_row, start_column=19, end_row=start_row, end_column=34)

    # Insert the filtered data into the Excel sheet and apply formatting
    for row_num, row_data in enumerate(filtered_data, start=start_row + 1):
        max_length_commentary = 0  # Initialize max length for Weekly Commentary text-based height adjustment
        max_length_next7days = 0  # Initialize max length for Next 7 Days text-based height adjustment
        line_breaks_commentary = 1  # Start with a base line count (1 line) for Weekly Commentary
        line_breaks_next7days = 1  # Start with a base line count (1 line) for Next 7 Days

        # Insert Channel Group in Column A and Campaign Group in Column B
        worksheet.cell(row=row_num, column=1, value=row_data[0])  # Channel Group
        worksheet.cell(row=row_num, column=2, value=row_data[1])  # Campaign Group

        for col_num, value in enumerate(row_data, start=start_column):
            if col_num == 3:  # Column C is index 3 when start_column is 1 (Weekly Commentary)
                worksheet.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=18)
                top_left_cell = worksheet.cell(row=row_num, column=3)
                top_left_cell.value = value
                top_left_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

                # Calculate the text length and number of line breaks
                text_length = len(value) if isinstance(value, str) else 0
                line_breaks_commentary += value.count('\n') if isinstance(value, str) else 0
                max_length_commentary = max(max_length_commentary, text_length)

                # Replace "." followed by whitespace with ".\n" to create a new line without leading space
                if isinstance(value, str):
                    top_left_cell.value = value.replace('. ', '.\n').replace(' .', '\n.')

            if col_num == 4:  # "Next 7 Days" is in column 4 (index 4 when start_column is 1)
                worksheet.merge_cells(start_row=row_num, start_column=19, end_row=row_num, end_column=34)
                top_left_cell = worksheet.cell(row=row_num, column=19)
                top_left_cell.value = value
                top_left_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

                # Calculate the text length and number of line breaks for Next 7 Days
                text_length = len(value) if isinstance(value, str) else 0
                line_breaks_next7days += value.count('\n') if isinstance(value, str) else 0
                max_length_next7days = max(max_length_next7days, text_length)

                # Replace "." followed by whitespace with ".\n" for Next 7 Days
                if isinstance(value, str):
                    top_left_cell.value = value.replace('. ', '.\n').replace(' .', '\n.')

        # Adjust row height dynamically based on text length and line breaks for both sections
        max_text_length = max(max_length_commentary, max_length_next7days)
        max_line_breaks = max(line_breaks_commentary, line_breaks_next7days)
        worksheet.row_dimensions[row_num].height = max(15, 15 + (max_text_length // 50) * 10 + max_line_breaks * 10)

        # Apply borders only to the relevant columns (A to AH) and avoid beyond
        for col_num in range(1, 35):  # Columns A to AH (1 to 34)
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.border = thin_border


    # Print the confirmation for each worksheet
    print(f"Formatted and filtered data from Google Sheets (Year: {iso_year}, ISO Week: {iso_week}) has been written to {excel_file_path}, starting at row {start_row}, column {start_column} in the '{worksheet_name}' worksheet.")

# Insert commentary into the "EOW" worksheet
insert_commentary("EOW")

# Insert commentary into the "WTD" worksheet
insert_commentary("WTD")

# Save the final Excel workbook with Google Sheets commentary included and formatted
workbook.save(excel_file_path)


# -------- Copy Worksheets from Another Workbook -------- #

# Load the source workbook
source_workbook = load_workbook('ga4-reports/input/xChannelGrowthMarketing_Read_Me.xlsx')

# Function to copy a worksheet manually
def copy_worksheet(source_sheet, target_workbook, new_sheet_name):
    new_sheet = target_workbook.create_sheet(title=new_sheet_name)
    
    # Copy cells and their values
    for row in source_sheet.iter_rows():
        for cell in row:
            new_cell = new_sheet[cell.coordinate]
            new_cell.value = cell.value
            
            # Copy cell style components individually, excluding color fill
            if cell.has_style:
                new_cell.font = Font(
                    name=cell.font.name,
                    size=cell.font.size,
                    bold=cell.font.bold,
                    italic=cell.font.italic,
                    vertAlign=cell.font.vertAlign,
                    underline=cell.font.underline,
                    strike=cell.font.strike,
                    color=cell.font.color
                )
                new_cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal,
                    vertical=cell.alignment.vertical,
                    text_rotation=cell.alignment.textRotation,
                    wrap_text=cell.alignment.wrap_text,
                    shrink_to_fit=cell.alignment.shrinkToFit,
                    indent=cell.alignment.indent
                )
                new_cell.border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom,
                    diagonal=cell.border.diagonal,
                    diagonal_direction=cell.border.diagonal_direction,
                    outline=cell.border.outline,
                    vertical=cell.border.vertical,
                    horizontal=cell.border.horizontal
                )
                new_cell.number_format = cell.number_format

    # Copy merged cell ranges
    for merged_range in source_sheet.merged_cells.ranges:
        new_sheet.merge_cells(str(merged_range))

    # Copy column dimensions
    for col in source_sheet.column_dimensions:
        new_sheet.column_dimensions[col] = source_sheet.column_dimensions[col]

    # Copy row dimensions
    for row_dim in source_sheet.row_dimensions:
        new_sheet.row_dimensions[row_dim] = source_sheet.row_dimensions[row_dim]

# Copy the "Report_name" and "Read_me" worksheets
copy_worksheet(source_workbook["Report_name"], workbook, "Report_name")
copy_worksheet(source_workbook["Read_me"], workbook, "Read_me")

# Move the copied sheets to the beginning of the workbook (before "WTD")
workbook._sheets.insert(0, workbook._sheets.pop(workbook._sheets.index(workbook["Report_name"])))
workbook._sheets.insert(1, workbook._sheets.pop(workbook._sheets.index(workbook["Read_me"])))

# -------- Update the "Report_name" and "Read_me" Worksheets -------- #

def format_worksheet(ws, max_row=200, max_col=50):
    # Iterate over a large range of cells in the worksheet and apply white fill and remove borders
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            cell.border = Border(left=Side(border_style=None),
                                 right=Side(border_style=None),
                                 top=Side(border_style=None),
                                 bottom=Side(border_style=None))

# Apply formatting to "Report_name" worksheet
report_name_sheet = workbook["Report_name"]
format_worksheet(report_name_sheet)

# Apply formatting to "Read_me" worksheet
read_me_sheet = workbook["Read_me"]
format_worksheet(read_me_sheet)

# -------- Specific Formatting for "Report_name" Worksheet -------- #

# Apply color fill to entire row 1
for col in range(1, report_name_sheet.max_column + 1):
    cell = report_name_sheet.cell(row=1, column=col)
    cell.fill = PatternFill(start_color="024991", end_color="024991", fill_type="solid")

# Apply color fill to rows starting from row 8 up to max_row
for row in range(8, 200 + 1):
    for col in range(1, report_name_sheet.max_column + 1):
        cell = report_name_sheet.cell(row=row, column=col)
        cell.fill = PatternFill(start_color="024991", end_color="024991", fill_type="solid")

# -------- Insert Image into "Report_name" Worksheet -------- #

# Load and insert the image into column B, row 3
from openpyxl.drawing.image import Image

image_path = 'ga4-reports/input/Picture1.png'
img = Image(image_path)

# Place the image at column B, row 3
report_name_sheet.add_image(img, 'B3')

# -------- Update the "Report_name" Worksheet -------- #

# Retrieve the last report date from the "WTD" worksheet (cell A2)
wtd_date_range = workbook["WTD"]['A2'].value
wtd_end_date_str = wtd_date_range.split(' to ')[1]  # Get the end date
wtd_end_date = datetime.strptime(wtd_end_date_str, '%m/%d/%Y').strftime('%B %d, %Y')  # Format as "Month Day, Year"

# Insert the last report date into cell B8 of the "Report_name" worksheet
report_name_sheet['B8'] = f"{report_name_sheet['B8'].value} {wtd_end_date}"
report_name_sheet['B8'].alignment = Alignment(horizontal='left')

# -------- Color Text in Cell B2 -------- #

# Color the text in cell B2 with Aptos Narrow font
cell_b2 = report_name_sheet['B2']
cell_b2.font = Font(name="Aptos Narrow", color="D65488", size=72, bold=True)  # Apply font, color, and size

# Save the final Excel workbook with the new worksheets included and formatted
workbook.save(excel_file_path)